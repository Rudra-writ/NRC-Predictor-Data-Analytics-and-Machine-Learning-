from django.shortcuts import render
import openpyxl
import pandas as pd
import math
from django.conf import settings
from NRCpredictor.models import AITableOne, AITrainerData, AITrainerDataScaled, accuracyTolerance, labourMaster, labourMasterAggregate, masterAggregate, materialMaster, materialMasterAggregate, completionRate, purchaseYearlyTotal, purchaseAggregate, masterYearlyTotal, requiredTable, graphData, yearlyGraphData
import sqlalchemy
import datetime as  dt
import json
from datetime import date
from django.contrib import messages
from plotly.offline import plot
from plotly.graph_objects import Scatter
from plotly.graph_objects import Layout
import numpy as np
from datetime import date
from django.http import HttpResponseRedirect
from django.http import HttpResponse

from operator import index
from keras.models import Sequential
from keras.layers import Dense
from keras.layers import Dropout
from sklearn import preprocessing
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import numpy as np
import keras
from keras.callbacks import TensorBoard
from matplotlib import pyplot
from keras.optimizers import adam_v2
from sklearn.model_selection import train_test_split
from keras.callbacks import EarlyStopping
from datetime import date
import pandas as pd 
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import LabelEncoder, RobustScaler
from keras.models import load_model
import statsmodels
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.arima.model import ARIMA
import statsmodels.api as sm
from pandas.tseries.offsets import DateOffset


import os

def first_page(request):
  if( request.method == 'POST'):
       return render(request, 'first_page.html')
  else:
    return render(request, 'first_page.html')

def index(request):
    Systemcode_lists=  ['S1PR-E01-TPU-SEN', 'GRND-EN1-MRO', 'S1PR-E02-3PB-MIX', 'GRND-HIL-MRO', 'S2VE-SEP-REC', 'BDSM-BDX-BJS', 'GRND-SLT', 'FCLY-WS1-CON', 'S3AV-NAV-INS', 'S1PR-E02-2TP-FUP', 'GRND-STR-FRG', 'GRND-HY1-CTN-MRO', 'S1PR-E02-0GA-OXL', 'S1AV-EUS-INT-PWR', 'BDSM-BDX-RAF-NAM', 'S1VE-TCR-SNS-PTS', 'C1AZ', 'S1PR-E01-TPU-BOV', 'S2VE-PLS', 'S1AV-NDE-N01', 'GRND-SLT-OST-PMD', 'GRND-FUR', 'GRND-STR-SEP', 'S2VE-RCS', 'S1PR-E02-0GA-SST', 'S1PR-E02-1EC-TCA-RDF', 'GRND-SHK-MRO', 'S1PR-E02-4CU-410', 'S3AV', 'GRND-INJ', 'S1AV', 'BDSM-BDX-EUC-ETP', 'C1AZ-PAD-GAS', 'S1PR-E02-0GA-FUL', 'GRND-PHS-MRO', 'S2VE-SEP-PHR', 'GRND-STR-REL-MRO', 'S1VE-PRS-THP', 'S1PR-E02-3PB-RDF', 'FCLY-WS2-MAC', 'S3VE-FRG-USH', 'S1VE-TFM-VAC-USV', 'S1PR-E02-0GA-MTH', 'S1VE-TCR-ITP', 'S1VE-TCR-PMD-ASL', 'GRND-PST', 'GRND-STR-MRO', 'S3VE-KST-STR', 'S1VE-PRS-RLF', 'S1PR-E02-4CU-410-RDF', 'BDSM-BDX', 'S1PR-E02-TVC-ACT', 'S2VE-TCR-SNS-AUX', 'GRND-INT-MRO', 'S2VE-SST', 'S1VE-TFM', 'GRND-PD2', 'FCLY-WS2-CLN', 'C2FR', 'GRND-TAS-SG2', 'GRND-HY1-CLN', 'S1VE-TFM-REL', 'BDSM-BDX-EMS', 'BDSM-BDX-ESA-PD2', 'FCLY-WS2-CON', 'S1VE', 'S2PR-E02-NEX', 'GRND-TPU-STR-MRO', 'GRND-STR-VE1-MRO', 'FDPT-EMS', 'S1AV-FTS', 'S1PR-E02-1EC-RDF', 'C1AZ-OPS', 'S3VE-FRG', 'GRND-HPH-MRO', 'S1VE-SEP-PHR', 'S1PR-E02-4CU-407-RDF', 'GRND-NM1-MRO', 'FCLY-WSP', 'C5DE', 'S3PR-K01-TCA-TCH', 'C1AZ-LOG', 'S3AV-NAV', 'FCLY-WS1-PPE', 'GRND-STR-COM', 'BDSM-BDX-ESA-CST-OP3', 'S1VE-PRS-RLF-ACT', 'BDSM-SAL-AST', 'S1PR-E01-TPU-GGU', 'GRND-PDT', 'S1PR-E02-4CU-407', 'S2VE-TCR-CNE', 'FCLY-WS2-TST', 'C1AZ-PAD-LOX', 'S1PR-E02-2TP-TUM-RDF', 'FDPT-NMO-UM2', 'S1PR-E01-TPU-SSS', 'S2VE-TCR-SNS-PTS', 'BDSM-BDX-ESA-CST', 'S1PR-E02-3PB-IGN', 'FDPT-ESA-CST-FSD', 'S1PR', 'BDSM-BDX-NMO', 'S3VE-FRG-SEP-REC', 'S3VE-FRG-SEP-PHR', 'Management', 'S1PR-E02-0GA', 'S3VE-KST-TNK-PMD-FLR', 'FCLY-WS2-MAC-MRO', 'GRND-BST', 'C1AZ-PAD-LN2', 'S2VE-PRS-THP-USG', 'S2AV-WRG', 'GRND-TS1-HTS', 'S1VE-TCR', 'GRND-ELC', 'FCLY-MCC-MCS', 'S1PR-E02-1EC-TCA', 'S1AV-PWR-BAT', 'GRND-PHS', 'GRND-TAS-SG1', 'S1PR-E01-TPU-TUR', 'S2VE-TCR-ITP', 'GRND-3DP-MRO', 'GRND-TS1-HTS-TB1', 'BDSM-BDX-F9X', 'GRND-TS1-HTS-MRO', 'S2PR-E02', 'S3PR-K01-TCA', 'GRND-VAC', 'S1PR-E02-4CU-405-RDF', 'BDSM-BDX-EUC-SCI', 'C1AZ-SLV', 'BDSM-BDX-MCS', 'S1PR-E02-3PB-TCA-RDF', 'S1PR-E02-4CU-408-RDF', 'FCLY-WSP-PPE', 'S2VE-TCR-PMD', 'S1PR-E01-TPU-IGN', 'S1AV-EUS', 'GRND-INS', 'S1VE-RVY', 'GRND-STR-VE1', 'S1PR-E02-RDF', 'GRND-TPU-PTR-MRO', 'S1VE-TFM-STR', 'IT', 'C1AZ-PAD-IXF-FSP', 'S1PR-E02-3PB-MIX-RDF', 'S2AV-FTS', 'C1AZ-FLY', 'S1PR-E01', 'GRND-TVC-LDB', 'FDPT-KSP', 'S1PR-E02-4CU-COX', 'S1AV-EUS-GRD-PWR', 'S2VE-PRS-THP-SNS', 'S1VE-TCR-SNS-LVL', 'GRND-TAS-SG1-MRO', 'S3AV-PWR', 'S2VE', 'S1PR-E02', 'GRND-TPU-TRG-MRO', 'GRND-TPU-MRO', 'C4CA', 'S1PR-E02-AUX', 'S3PR-K01-TRV', 'S1VE-SEP-REC', 'S1PR-E02-4CU-409', 'S1VE-PRS-RLF-300', 'C1AZ-PAD-LPS', 'S1PR-E02-0GA-SNS', 'S2AV-COM', 'S1PR-E02-1EC', 'S3VE', 'S1AV-RTS', 'S1AV-EUS-GRD', 'GRND-HIL-SCE', 'S1VE-TCR-PMD-FLR', 'S2VE-TCR-PMD-ASL', 'S3PR', 'GRND-INT', 'GRND-HFR', 'S1PR-E02-4CU-FTS', 'S3VE-RCS-USG', 'FCLY-WS2-PPE', 'GRND-BST-MRO', 'S2PR-E02-AUX', 'S2AV-EUS', 'S2VE-RCS-USG', 'S1AV-PWR', 'S1VE-TCR-ASK', 'S1PR-E02-4CU-411', 'S2VE-TCR-BSF', 'S2VE-PRS-THP-HES', 'S3VE-KST-TNK-PMD-AVX', 'GRND-SHK', 'GRND-STR-REL', 'S3VE-FRG-SEP-ACT', 'GRND-STR-COM-MRO', 'S3AV-COM', 'GRND-TPU', 'S3AV-NDE', 'S2AV', 'GRND-INS-MRO', 'S1PR-E02-1EC-MIX-RDF', 'S3VE-KST-SNS', 'GRND-TAS-SG2-MRO', 'S1PR-E01-TPU-SEL', 'S1VE-TCR-SNS', 'C1AZ-DGS', 'FCLY-WS2-PRO', 'S3AV-WRG', 'S1AV-NDE', 'S1VE-TCR-FSK', 'S1PR-E01-TCA', 'C1NO-PAD', 'S2PR', 'S1VE-PRS-RLF-200 ', 'S1AV-WRG', 'S1PR-E02-TVC-MTH', 'C1NO-LIC', 'S3VE-SEP-REC', 'BDSM-BDX-BAY', 'GRND-3DP', 'S1AV-EUS-INT', 'GRND-TAS-SWG-MRO', 'BDSM-MKT', 'GRND-SNS-PSS', 'S3VE-KST-USS', 'GRND-HPH', 'FDPT-ESA', 'BDSM-BDX-RAF-CCT', 'FCLY-IT-ITI ', 'GRND-TS1-VTS-IST', 'FCLY-SSC', 'GRND-SLT-OST-PMD-FLR', 'S3PR-K01-TVC', 'BDSM-BDX-RAF', 'GRND-TS1', 'S3VE-FRG-ITP', 'BDSM-BDX-ESA-FAM', 'S1VE-SEP-ACT', 'S1PR-E02-2TP', 'S3AV-EUS', 'GRND-TS1-VTS-MRO', 'GRND-TS1-HTS-TB1-PFT', 'GRND-VAC-MRO', 'S2VE-SEP-ACT', 'S1VE-PLS-USR', 'BDSM-MKT-SMX', 'BDSM-BDX-ESA-CST-OP2', 'S2VE-TCR-TFM', 'S1VE-PRS-THP-HES', 'S2AV-NDE', 'S1VE-PLS-USC', 'BDSM-MKT-WEB', 'BDSM-BDX-EUC', 'S1PR-E02-4CU-408', 'S1PR-E02-0GA-GDT', 'S1VE-TCR-TSF', 'FCLY-IT-USR ', 'BDSM-BDX-EUC-LCL', 'GRND-CRS', 'C1AZ-PAD', 'GRND-HY1-HST', 'GRND-STR-SEP-MRO', 'S1PR-E02-0GA-OXL-APD', 'C6OM', 'S3VE-KST-HES', 'S3VE-KST-USG', 'GRND-SNS-MRO', 'FDPT-NMO', 'S2VE-TCR-STR', 'S1PR-E02-4CU-402-RDF', 'BDSM-SAL-OIT', 'S3AV-NAV-TSO', 'S1VE-SEP', 'BDSM-MKT-SYM', 'S1PR-E02-1EC-IGN', 'C1NO-FLY', 'BDSM-SAL-PIM', 'S1PR-E02-4CU-405', 'S3VE-KST-TNK', 'S2VE-PRS-THP', 'S2VE-TCR-SNS', 'S2VE-PLS-MCV', 'FCLY-WS2', 'GRND-HY1-CTN', 'S1PR-E02-AUX-RDF', 'FCLY-C1 ', 'FDPT-AWO', 'S3VE-KST-TNK-PMD-ASL', 'GRND-TS1-VTS-IEQ', 'S1PR-E02-0GA-HTS', 'S1PR-E02-3PB', 'BDSM-SAL', 'S1VE-PLS-MCV', 'GRND-TPU-PTR', 'GRND-STR-KST-MRO', 'FDPT-MCS', 'S1AV-COM', 'S1PR-E02-3PB-IGN-RDF', 'S3PR-K01-TCA-INJ', 'GRND-TVC-MRO', 'S1PR-E02-4CU-402', 'S3VE-FRG-SNS', 'GRND-HFR-MRO', 'GRND-TPU-STR', 'S1PR-E01-TPU-CLT', 'GRND-TPU-TRG', 'FDPT-ESA-LCA', 'S1VE-TFM-SST', 'S1PR-E02-2TP-OXP', 'S1PR-E02-0GA-PVC', 'S1PR-E02-TVC-PWR', 'C2FR-FLY', 'FCLY-WSP-OFF', 'C1NO', 'S1PR-E02-4CU-406', 'S2VE-TCR-PMD-AVX', 'FCLY-WS2-REL', 'S1PR-E02-0GA-IGN', 'S1VE-TFM-USH', 'GRND-TS1-VTS', 'GRND-PST-MRO', 'S1VE-PRS-THP-USG', 'S1AV-RVY', 'S1VE-TFM-VAC', 'BDSM-BDX-ESA-CST-FP1', 'S3AV-RFS', 'C2FR-GRY', 'BDSM-BDX-NMO-EL3', 'S1PR-E02-4CU-401', 'GRND-NM1', 'GRND-TS1-HTX-MRO', 'FCLY-TC1', 'S1VE-TCR-SNS-AUX', 'BDSM-SAL-CLS', 'BDSM-BDX-RAF-CBP', 'BDSM-SAL-RAY', 'S2VE-PRS-REG', 'S3VE-DSP', 'C1AZ-PAD-WDS', 'S2AV-PWR', 'C1AZ-PAD-AIF', 'S1VE-PRS-THP-SNS', 'S2VE-TCR-TSF', 'S1PR-E02-ECU', 'FCLY-WSP-PRO', 'GRND-HY1-MRO', 'FDPT-ESA-CST-OSD', 'FCLY-IT ', 'GRND-TVC', 'GRND-EN1', 'BDSM-BDX-RAF-TPB', 'S3VE-FRG-FLR', 'GRND-HY1', 'S1PR-E02-1EC-IGN-RDF', 'S1PR-E02-TVC-GIM', 'S1PR-E02-0GA-PUL', 'S3VE-SEP-ACT', 'GRND-TS1-HTS-TB1-PWP', 'BDSM-BDX-ESA', 'S1PR-E02-0GA-GEJ', 'C7UK', 'S1PR-E02-0GA-PRS', 'S1PR-E01-TPU', 'S2VE-TCR-SNS-LVL', 'C1AZ-LCC', 'S1PR-E02-4CU-403-RDF', 'C1AZ-PAD-IXF', 'S1VE-PRS', 'S1VE-TCR-PMD', 'S1VE-PRS-REG', 'BDSM-BDX-KSP', 'GRND-HIL', 'C1AZ-SEA', 'S3VE-KST-ITP', 'GRND-SLT-IST', 'S1VE-PLS', 'S3VE-FRG-SEP', 'S3VE-KST', 'S3VE-SEP-PHR', 'FCLY-WS2-ELC', 'GRND-TAS-MRO', 'S2VE-PRS-RLF', 'C1', 'S3AV-FCS', 'S2VE-PLS-USC', 'GRND-TS1-HTX', 'GRND-TS1-MRO', 'S2VE-TCR', 'GRND-PD2-MRO', 'BDSM-BDX-ESA-CST-OP1', 'GRND-INJ-MRO', 'FCLY-WSP-CON', 'S3VE-SEP',  'GRND-STR', 'GRND-FUR-MRO', 'S1VE-TCR-STR', 'S2VE-PLS-USR', 'GRND-STR-KST', 'GRND-SLT-HEX', 'GRND-TAS', 'BDSM-BDX-ESA-LCA', 'GRND-SNS', 'GRND-TAS-SWG', 'GRND-PDT-MRO', 'C2FR-OPS', 'C1AZ-PAD-ERC', 'S1PR-E02-3PB-TCA', 'S1VE-TCR-FSK-USH', 'BDSM-MKT-EVT', 'GRND-STR-FRG-MRO', 'C1NO-IXF', 'S1PR-E02-4CU-403', 'S3VE-RCS', 'FCLY-MCC', 'GRND-HY1-CLN-MRO', 'BDSM-BDX-AIC', 'GRND-HY1-HST-MRO', 'S3VE-FRG-STR', 'S3VE-KST-SAB', 'FDPT-ESA-CST', 'BDSM-SAL-PUG', 'FDPT-AIC', 'S1AV-EUS-INT-COM', 'GRND-SLT-OST-PMD-AVX', 'S3PR-K01-TCA-IGN', 'C1AZ-GRY', 'S1PR-E02-0GA-GDT-RDF', 'C1AZ-PAD-UM1', 'S3AV-NAV-SIM', 'S1PR-E01-TPU-PMP', 'S2VE-PRS', 'S1VE-TCR-PMD-AVX', 'FDPT-ESA-NST', 'BDSM-SAL-OSE', 'GRND-SLT-OST', 'BDSM-BDX-RAF-REM', 'S1PR-E02-4CU-404', 'GRND-NM1-DAQ', 'S2VE-TCR-PMD-FLR', 'BDSM-BDX-EUC-HEU', 'S1PR-E02-4CU', 'FCLY-WS2-OFF', 'S3VE-KST-TNK-PMD', 'S2VE-SEP', 'FCLY-WS1', 'GRND-SLT-OST-PMD-ASL', 'S1AV-EUS-GRD-COM', 'S3PR-K01', 'S1PR-E02-1EC-MIX', 'S1PR-E02-SLE', 'S1PR-E02-TVC', 'BDSM-BDX-NMO-UM2', 'S1VE-TFM-SST-USS', 'S1PR-E02-2TP-TUM']
    
    if "POST" == request.method:
      df_Labour_master = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      df_Labour_aggregate = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      df_Material_master = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])      
      df_Material_aggregate = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      df_purchaseOrder = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total' ] )
      df1 = pd.DataFrame()
      df2 = pd.DataFrame()
      df_completionRate = pd.DataFrame(columns = ['system_code', 'completion_rate'])
      df_DB_format = pd.DataFrame(columns = ['system_code', 'total_cost', 'Date'  ] )
      df_DB_aggregate = pd.DataFrame(columns = ['system_code', 'total'])
      df_DB_yearlytotal= pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
    #HANDLING DATA FROM THE '''''LABOUR SHEET''''' OF THE CONTROLLING_ACC.NOV  EXCEL FILE  
      if(request.POST.get('upload')):
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file, data_only= True)
        worksheet1 = wb["Labour"]
        df = pd.DataFrame(worksheet1.values)
        df_row = pd.DataFrame()
        df_Labour = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
    
        columns_no = len(df.columns)-1
        df.fillna(0, inplace = True)
        for sys_co in Systemcode_lists:
          if sys_co in df.iloc[:,0].values:
            complete_year = math.floor(columns_no/12)
            incomplete_year = columns_no % 12   
            n = 1 
            m = 0
            init_year= 2019
            df_row[0:0]
            df_row = df.loc[df.iloc[:,0]== sys_co]
            df_row.reset_index(drop = True, inplace = True)
            for x in range(0, complete_year):
                    
                    df_Labour.at[0,'system_code'] = sys_co
                    df_Labour.at[0,'year'] = init_year
                    df_Labour['january'] = abs(df_row.iloc[0,0+n] - df_row.iloc[0,0+m]) if x != 0 else df_row.iloc[0,0+n]
                    df_Labour['february'] = abs(df_row.iloc[0,1+n] - df_row.iloc[0,0+n])
                    df_Labour['march'] = abs( df_row.iloc[0,2+n] - df_row.iloc[0,1+n])
                    df_Labour['april'] = abs(df_row.iloc[0,3+n] - df_row.iloc[0,2+n])
                    df_Labour['may'] = abs(df_row.iloc[0,4+n] - df_row.iloc[0,3+n])
                    df_Labour['june'] = abs(df_row.iloc[0,5+n] - df_row.iloc[0,4+n])
                    df_Labour['july'] = abs(df_row.iloc[0,6+n] - df_row.iloc[0,5+n])
                    df_Labour['august'] = abs(df_row.iloc[0,7+n] - df_row.iloc[0,6+n])
                    df_Labour['september'] = abs(df_row.iloc[0,8+n] - df_row.iloc[0,7+n])
                    df_Labour['october'] = abs(df_row.iloc[0,9+n] - df_row.iloc[0,8+n])
                    df_Labour['november'] = abs(df_row.iloc[0,10+n] - df_row.iloc[0,9+n])
                    df_Labour['december'] = abs(df_row.iloc[0,11+n] - df_row.iloc[0,10+n])
                    df_Labour['total'] = df_Labour.loc[:,'january':'december'].sum(axis=1)
                    df_Labour_master = df_Labour_master.append(df_Labour)
                    df_Labour = df_Labour[0:0]
                    n=n + 12
                    m = m + 12
                    init_year = init_year + 1
            final_n = n
            final_year = init_year  
            final_m = m        
            df_Labour.at[0,'system_code'] = sys_co
            df_Labour.at[0,'year'] = final_year 
            df_Labour['january'] = abs(df_row.iloc[0,0+final_n] - df_row.iloc[0,0+final_m]) if incomplete_year >= 1 else 0
            df_Labour['february'] = abs(df_row.iloc[0,1+final_n] - df_row.iloc[0,0+final_n]) if incomplete_year >= 2 else 0
            df_Labour['march'] = abs(df_row.iloc[0,2+final_n] - df_row.iloc[0,1+final_n]) if incomplete_year >= 3 else 0
            df_Labour['april'] = abs(df_row.iloc[0,3+final_n] - df_row.iloc[0,2+final_n]) if incomplete_year >= 4 else 0
            df_Labour['may'] = abs(df_row.iloc[0,4+final_n] - df_row.iloc[0,3+final_n]) if incomplete_year >= 5 else 0
            df_Labour['june'] = abs(df_row.iloc[0,5+final_n] - df_row.iloc[0,4+final_n]) if incomplete_year >= 6 else 0
            df_Labour['july'] = abs(df_row.iloc[0,6+final_n] - df_row.iloc[0,5+final_n]) if incomplete_year >= 7 else 0
            df_Labour['august'] = abs(df_row.iloc[0,7+final_n] - df_row.iloc[0,6+final_n]) if incomplete_year >= 8 else 0
            df_Labour['september'] = abs(df_row.iloc[0,8+final_n] - df_row.iloc[0,7+final_n]) if incomplete_year >=  9 else 0
            df_Labour['october'] = abs(df_row.iloc[0,9+final_n] - df_row.iloc[0,8+final_n]) if incomplete_year >= 10 else 0
            df_Labour['november'] = abs(df_row.iloc[0,10+final_n] - df_row.iloc[0,9+final_n]) if incomplete_year >= 11 else 0
            df_Labour['december'] = abs(df_row.iloc[0,11+final_n] - df_row.iloc[0,10+final_n]) if incomplete_year == 12 else 0
            df_Labour['total'] = df_Labour.loc[:, 'january':'november'].sum(axis=1)
            df_Labour_master = df_Labour_master.append(df_Labour)
            df_Labour = df_Labour[0:0]
                      
        df_Labour_master = df_Labour_master[df_Labour_master['system_code'].notna()]
        df_Labour_master.reset_index(drop =True, inplace = True)     #master table for Labours data
        
        #Migrate the dataframe to database table labour_master
        model= labourMaster.objects.all()
        model.delete()
        for index, row in df_Labour_master.iterrows():
               model = labourMaster()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january = round(abs(row['january']),2)
               model.february = round(abs(row['february']),2)
               model.march = round(abs(row['march']),2)
               model.april = round(abs(row['april']),2)
               model.may = round(abs(row['may']),2)
               model.june = round(abs(row['june']),2)
               model.july = round(abs(row['july']),2)
               model.august = round(abs(row['august']),2)
               model.september = round(abs(row['september']),2)
               model.october = round(abs(row['october']),2)
               model.november = round(abs(row['november']),2)
               model.december = round(abs(row['december']),2)
               model.total = round(row['total'],2)
               model.save()
       
        
        #labourmaster.objects.bulk_create(labourmaster(**vals) for vals in df_Labour_master.to_dict())

        print(df_Labour_master.shape)

        df_Labour_aggregate = df_Labour_master.groupby( ["system_code"], as_index = False).sum()
        df_Labour_aggregate= df_Labour_aggregate.drop('year', axis = 1) 
        df_Labour_aggregate = df_Labour_aggregate.reset_index(drop=True) #Table for aggregate of Labour cost of each system code over all the years
        
        #Migrate the dataframe to database table labour_master_aggregate
        model= labourMasterAggregate.objects.all()
        model.delete()
        for index, row in df_Labour_aggregate.iterrows():
               model = labourMasterAggregate()
               model.system_code = row['system_code']
              
               model.total = round(row['total'],2)
               model.save()


          
    #HANDLING DATA FROM THE '''''MATERIAL SHEET''''' OF THE CONTROLLING_ACC.NOV  EXCEL FILE
        worksheet2 = wb["Material"]
        df=df[0:0]
        df = pd.DataFrame(worksheet2.values)
        df_row2 = pd.DataFrame()
        df_Material = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
    
        columns_no = len(df.columns)-1
        df.fillna(0, inplace = True)
        for sys_co in Systemcode_lists:
          if sys_co in df.iloc[:,0].values:
            complete_year = math.floor(columns_no/12)
            incomplete_year = columns_no % 12
            n = 1 
            m = 0
            init_year= 2019
            df_row2[0:0]
            df_row2 = df.loc[df.iloc[:,0]== sys_co]
            df_row2.reset_index(drop = True, inplace = True)
            for x in range(0, complete_year):
                    
                    df_Material.at[0,'system_code'] = sys_co
                    df_Material.at[0,'year'] = init_year
                    df_Material['january'] = abs(df_row2.iloc[0,0+n] - df_row2.iloc[0,0+m]) if x != 0 else df_row2.iloc[0,0+n]
                    df_Material['february'] = abs(df_row2.iloc[0,1+n] - df_row2.iloc[0,0+n])
                    df_Material['march'] = abs(df_row2.iloc[0,2+n] - df_row2.iloc[0,1+n])
                    df_Material['april'] = abs(df_row2.iloc[0,3+n] - df_row2.iloc[0,2+n])
                    df_Material['may'] = abs(df_row2.iloc[0,4+n] - df_row2.iloc[0,3+n])
                    df_Material['june'] = abs(df_row2.iloc[0,5+n] - df_row2.iloc[0,4+n])
                    df_Material['july'] = abs(df_row2.iloc[0,6+n] - df_row2.iloc[0,5+n])
                    df_Material['august'] = abs(df_row2.iloc[0,7+n] - df_row2.iloc[0,6+n])
                    df_Material['september'] = abs(df_row2.iloc[0,8+n] - df_row2.iloc[0,7+n])
                    df_Material['october'] = abs(df_row2.iloc[0,9+n] - df_row2.iloc[0,8+n])
                    df_Material['november'] = abs(df_row2.iloc[0,10+n] - df_row2.iloc[0,9+n])
                    df_Material['december'] = abs(df_row2.iloc[0,11+n] - df_row2.iloc[0,10+n])
                    df_Material['total'] = df_Material.loc[:,'january':'december'].sum(axis=1)
                    df_Material_master = df_Material_master.append(df_Material)
                    df_Material = df_Material[0:0]
                    n=n + 12
                    m = m + 12
                    init_year = init_year + 1
            final_n = n
            final_year = init_year
            final_m = m
            
            df_Material.at[0,'system_code'] = sys_co
            df_Material.at[0,'year'] = final_year 
            df_Material['january'] = abs(df_row2.iloc[0,0+final_n] - df_row2.iloc[0,0+final_m])  if incomplete_year >= 1 else 0
            df_Material['february'] = abs(df_row2.iloc[0,1+final_n] - df_row2.iloc[0,0+final_n]) if incomplete_year >= 2 else 0
            df_Material['march'] =  abs(df_row2.iloc[0,2+final_n] - df_row2.iloc[0,1+final_n]) if incomplete_year >= 3 else 0
            df_Material['april'] = abs(df_row2.iloc[0,3+final_n] - df_row2.iloc[0,2+final_n]) if incomplete_year >= 4 else 0
            df_Material['may'] = abs(df_row2.iloc[0,4+final_n] - df_row2.iloc[0,3+final_n]) if incomplete_year >= 5 else 0
            df_Material['june'] = abs(df_row2.iloc[0,5+final_n] - df_row2.iloc[0,4+final_n]) if incomplete_year >= 6 else 0
            df_Material['july'] = abs(df_row2.iloc[0,6+final_n] - df_row2.iloc[0,5+final_n]) if incomplete_year >= 7 else 0
            df_Material['august'] = abs(df_row2.iloc[0,7+final_n] - df_row2.iloc[0,6+final_n]) if incomplete_year >= 8 else 0
            df_Material['september'] = abs(df_row2.iloc[0,8+final_n] - df_row2.iloc[0,7+final_n]) if incomplete_year >=  9 else 0
            df_Material['october'] = abs(df_row2.iloc[0,9+final_n] - df_row2.iloc[0,8+final_n]) if incomplete_year >= 10 else 0
            df_Material['november'] =  abs(df_row2.iloc[0,10+final_n] - df_row2.iloc[0,9+final_n]) if incomplete_year >= 11 else 0
            df_Material['december'] = abs(df_row2.iloc[0,11+final_n] - df_row2.iloc[0,10+final_n]) if incomplete_year == 12 else 0
            df_Material['total'] = df_Material.loc[:, 'january':'november'].sum(axis=1)
            df_Material_master = df_Material_master.append(df_Material)
            df_Material = df_Material[0:0]         

      #Migrate the dataframe to database table material_master
        model= materialMaster.objects.all()
        model.delete()
        for index, row in df_Material_master.iterrows():
               model = materialMaster()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january = round(abs(row['january']),2)
               model.february = round(abs(row['february']),2)
               model.march = round(abs(row['march']),2)
               model.april = round(abs(row['april']),2)
               model.may = round(abs(row['may']),2)
               model.june = round(abs(row['june']),2)
               model.july = round(abs(row['july']),2)
               model.august = round(abs(row['august']),2)
               model.september = round(abs(row['september']),2)
               model.october = round(abs(row['october']),2)
               model.november = round(abs(row['november']),2)
               model.december = round(abs(row['december']),2)
               model.total = round(row['total'],2)
               model.save()

               
        df_Material_master = df_Material_master[df_Material_master['system_code'].notna()]
        df_Material_master.reset_index(drop =True, inplace = True)     #master table for Materials data
        df_Material_aggregate = df_Material_master.groupby( ["system_code"], as_index = False).sum()
        df_Material_aggregate= df_Material_aggregate.drop('year', axis = 1) 
        df_Material_aggregate = df_Material_aggregate.reset_index(drop=True) #Table for aggregate of Materials cost for each system code over all the years
        
        #Migrate the dataframe to database table material_master_aggregate
        model= materialMasterAggregate.objects.all()
        model.delete()
        for index, row in df_Material_aggregate.iterrows():
               model = materialMasterAggregate()
               model.system_code = row['system_code']
              
               model.total = round(row['total'],2)
               model.save()
      
        return render(request, 'index.html')
      
      elif (request.POST.get('upload3')):

        excel_file3 = request.FILES["excel_file3"]
        wb = openpyxl.load_workbook(excel_file3 , data_only= True)
        worksheet3 = wb["MAIN"]
        df = pd.DataFrame(worksheet3.values)
        df_row3 = pd.DataFrame()
        
        df1= df.iloc[7:479, 2:8]
       
        df1['system_code'] = df1.iloc[:,0].fillna('*') + df1.iloc[:,1].fillna('*') + df1.iloc[:,2].fillna('*') + df1.iloc[:,3].fillna('*') + df1.iloc[:,4].fillna('*') + df1.iloc[:,5].fillna('*')
        df1.drop(df1.columns[0:6], axis = 1, inplace = True)
        df1['system_code']= df1['system_code'].str.split(' ').str[0]
        df1['system_code'] = df1['system_code'].str.replace("*", "")
        df1['system_code'] = df1['system_code'].str.strip()
        
        df2 = df.iloc[7:479, 14:20].astype(str)
        df2['completion_rate'] = df2.iloc[:,0].fillna('*') + df2.iloc[:,1].fillna('*') + df2.iloc[:,2].fillna('*') + df2.iloc[:,3].fillna('*') + df2.iloc[:,4].fillna('*') + df2.iloc[:,5].fillna('*')
        df2.drop(df2.columns[0:6],axis=1,inplace=True)
        df2['completion_rate'] = df2['completion_rate'].str.replace("None", "")
        df2['completion_rate'] = df2['completion_rate'].str.replace("N/A", "0")
        df_completionRate['system_code'] = df1['system_code']
        df_completionRate['completion_rate'] = df2['completion_rate']
        df_completionRate.reset_index(drop = True, inplace = True)  #Data frame containing the Completion rates against system codes.
        print(df_completionRate.tail(25))
        
        #Migrate the dataframe to database table completion_rate
        model= completionRate.objects.all()
        model.delete()
        for index, row in df_completionRate.iterrows():
               model = completionRate()
               model.system_code = row['system_code']
              
               model.completion_rate = row['completion_rate']
               model.save()

        return render(request, 'index.html')

     #HANDLING DATA FROM THE '''''Incoming Goods SHEET''''' OF THE DB Search Results EXCEL FILE
      elif (request.POST.get('upload2')):
        excel_file2 = request.FILES["excel_file2"]
        wb = openpyxl.load_workbook(excel_file2, data_only=True)
        worksheet2 = wb["Incoming Goods"]
        df = pd.DataFrame(worksheet2.values)
        df_row3 = pd.DataFrame()
        df1 = df.iloc[1:, :]
        
        df_DB_format = df1.iloc[:,[3,2,8]]
        df_DB_format.rename(columns={3:'system_code', 2: 'total_cost', 8: 'Date' }, inplace = True)
        df_DB_format['Date'] = pd.to_datetime(df_DB_format['Date'], format ='%d/%m/%Y')
        df_DB_format.reset_index(drop = True, inplace=True)
        df_DB_format['total_cost'] = df_DB_format['total_cost'].astype(float)
     
        for axis, row in df_DB_format.iterrows():
          if row.system_code not in Systemcode_lists:
            df_DB_format.drop(axis, inplace = True)
        df_DB_format.reset_index(drop=True, inplace=True)
        df_DB_format = df_DB_format.drop_duplicates()
       
        for axis, row in df_DB_format.iterrows():
            df_purchaseOrder.at[axis,'system_code'] = row['system_code']
            df_purchaseOrder.at[axis,'year'] = row['Date'].year
            df_purchaseOrder.at[axis,'january'] = row['total_cost']if row['Date'].month == 1 else 0
            df_purchaseOrder.at[axis,'february'] = row['total_cost'] if row['Date'].month== 2 else 0
            df_purchaseOrder.at[axis,'march'] = row['total_cost'] if row['Date'].month == 3 else 0
            df_purchaseOrder.at[axis,'april'] = row['total_cost'] if row['Date'].month == 4 else 0
            df_purchaseOrder.at[axis,'may'] = row['total_cost'] if row['Date'].month == 5 else 0
            df_purchaseOrder.at[axis,'june'] = row['total_cost'] if row['Date'].month == 6 else 0
            df_purchaseOrder.at[axis,'july'] = row['total_cost'] if row['Date'].month == 7 else 0
            df_purchaseOrder.at[axis,'august'] = row['total_cost'] if row['Date'].month == 8 else 0
            df_purchaseOrder.at[axis,'september'] = row['total_cost'] if row['Date'].month == 9 else 0
            df_purchaseOrder.at[axis,'october'] = row['total_cost'] if row['Date'].month == 10 else 0
            df_purchaseOrder.at[axis,'november'] = row['total_cost'] if row['Date'].month == 11 else 0
            df_purchaseOrder.at[axis,'december'] = row['total_cost'] if row['Date'].month == 12 else 0
        df_purchaseOrder['total'] = df_purchaseOrder.loc[:,'january':'december'].sum(axis=1)

        df_purchaseOrder.sort_values(by=['system_code'], inplace= True) 
        laabels =  { 'january' : 'mean', 'february': 'mean', 'march': 'mean', 'april': 'mean', 'may': 'mean', 'june': 'mean', 'july': 'mean', 'august': 'mean', 'september': 'mean', 'october': 'mean', 'november': 'mean', 'december': 'mean', 'total': 'mean' } 
        df_purchaseOrder = df_purchaseOrder.groupby(['system_code', 'year'], as_index =False).agg(laabels)
        df_purchaseOrder.reset_index(drop = True, inplace = True) #To be used as A.I training dataframe
        df_purchaseOrder.sort_values(['system_code', 'year'], inplace = True)

        #Migrating df_purchaseOrder to ai_table_one
        model= AITableOne.objects.all()
        model.delete()
        for index, row in df_purchaseOrder.iterrows():
               model = AITableOne()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january = round(row['january'],2)
               model.february = round(row['february'],2)
               model.march = round(row['march'],2)
               model.april = round(row['april'],2)
               model.may = round(row['may'],2)
               model.june = round(row['june'],2)
               model.july = round(row['july'],2)
               model.august = round(row['august'],2)
               model.september = round(row['september'],2)
               model.october = round(row['october'],2)
               model.november = round(row['november'],2)
               model.december = round(row['december'],2)
               model.master_total = row['total']
               model.save()


        print(df_purchaseOrder.head(50))
        labels = {'january':'sum', 'february':'sum', 'march':'sum', 'april':'sum', 'may':'sum', 'june':'sum', 'july':'sum', 'august':'sum', 'september':'sum', 'october':'sum', 'november':'sum', 'december':'sum', 'total':'sum'}  
        df_DB_yearlytotal = df_purchaseOrder.groupby(["system_code", "year"], as_index = False).agg(labels) 
        df_DB_yearlytotal.reset_index(drop = True, inplace = True) # DataFrame containing the yearly aggregate of purchase order over all the months for purchase orders
        
        #Migrating df_DB_yearlytotal DataFrame to the purchase_yearly_total Table
        model= purchaseYearlyTotal.objects.all()
        model.delete()
        for index, row in df_DB_yearlytotal.iterrows():
               model = purchaseYearlyTotal()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january = round(row['january'],2)
               model.february = round(row['february'],2)
               model.march = round(row['march'],2)
               model.april = round(row['april'],2)
               model.may = round(row['may'],2)
               model.june = round(row['june'],2)
               model.july = round(row['july'],2)
               model.august = round(row['august'],2)
               model.september = round(row['september'],2)
               model.october = round(row['october'],2)
               model.november = round(row['november'],2)
               model.december = round(row['december'],2)
               model.total = round(row['total'],2)
               model.save()

        df_DB_aggregate = df_DB_yearlytotal.groupby(["system_code"], as_index = False).sum()
        #DataFrame containing the total of purchase orders over all the year till date
        df_DB_aggregate.drop(['year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december'], axis = 1, inplace = True)
        
        #Migrating df_DB_aggregate DataFrame to the purchase_aggregate Table
        model= purchaseAggregate.objects.all()
        model.delete()
        for index, row in df_DB_aggregate.iterrows():
               model = purchaseAggregate()
               model.system_code = row['system_code']
              
               model.total = round(row['total'],2)
               model.save()
        
        print(df_DB_aggregate.head(50))
        return render(request, 'index.html')
      
      #Whatever happens in this block is internal to this block
      elif(request.POST.get('analyze')):

        material_master = materialMaster.objects.all().values()
        labour_master = labourMaster.objects.all().values()
        purchase_yearly_total = purchaseYearlyTotal.objects.all().values()
        material_aggregate = materialMasterAggregate.objects.all().values()
        labour_aggregate = labourMasterAggregate.objects.all().values()
        purchase_aggregate = purchaseAggregate.objects.all().values()
        completion_rate = completionRate.objects.all().values()

        material_master_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        labour_master_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        purchase_yearly_total_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total']) 
        material_aggregate_df = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        labour_aggregate_df  = pd.DataFrame(columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        purchase_aggregate_df = pd.DataFrame(columns = [ 'system_code', 'total'])
        completion_rate_df = pd.DataFrame(columns = [ 'system_code', 'completion_rate'])

        material_aggregate_df = pd.DataFrame(material_aggregate)
        labour_aggregate_df = pd.DataFrame(labour_aggregate)
        purchase_aggregate_df = pd.DataFrame(purchase_aggregate)
        completion_rate_df = pd.DataFrame(completion_rate)
        material_master_df = pd.DataFrame(list(material_master))
        labour_master_df = pd.DataFrame(list(labour_master))
        purchase_yearly_total_df = pd.DataFrame(list(purchase_yearly_total))
        master_aggregate = pd.DataFrame()

        master_yearly_total_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])

        test_df = pd.DataFrame()

        required_df = pd.DataFrame()

        test_df = pd.merge(purchase_yearly_total_df,material_master_df, how = 'outer', on= ['system_code', 'year'])
        test_df.drop(['id_x', 'id_y', 'total_x', 'total_y'], axis = 1, inplace = True)
        test_df = test_df.merge(labour_master_df, how='outer', on = ['system_code', 'year'])
        test_df.drop(['id', 'total'], axis = 1, inplace = True)
        test_df.fillna(0, inplace= True)
        for index, row in test_df.iterrows():
          master_yearly_total_df.at[index, 'system_code'] = row['system_code']
          master_yearly_total_df.at[index, 'year'] = row['year']
          master_yearly_total_df.at[index, 'january'] = row['january_x'] + row['january_y'] + row['january']
          master_yearly_total_df.at[index, 'february'] = row['february_x'] + row['february_y'] + row['february']
          master_yearly_total_df.at[index, 'march'] = row['march_x'] + row['march_y'] + row['march']
          master_yearly_total_df.at[index, 'april'] = row['april_x'] + row['april_y'] + row['april']
          master_yearly_total_df.at[index, 'may'] = row['may_x'] + row['may_y'] + row['may']
          master_yearly_total_df.at[index, 'june'] = row['june_x'] + row['june_y'] + row['june']
          master_yearly_total_df.at[index, 'july'] = row['july_x'] + row['july_y'] + row['july']
          master_yearly_total_df.at[index, 'august'] = row['august_x'] + row['august_y'] + row['august']
          master_yearly_total_df.at[index, 'september'] = row['september_x'] + row['september_y'] + row['september']
          master_yearly_total_df.at[index, 'october'] = row['october_x'] + row['october_y'] + row['october']
          master_yearly_total_df.at[index, 'november'] = row['november_x'] + row['november_y'] + row['november']
          master_yearly_total_df.at[index, 'december'] = row['december_x'] + row['december_y'] + row['december']
        master_yearly_total_df['master_total'] = master_yearly_total_df.loc[:,'january':'december'].sum(axis=1) #DataFrame with monthly and yearly aggregate of material cost + labour cost + purchase order for each system code over individual year
        
        labels = { 'master_total':'sum'}
        
        master_aggregate = master_yearly_total_df.groupby(['system_code'], as_index = False).agg(labels) #DataFrame containing the yearly aggregate of  material cost + labour cost + purchase order for each system code over all years
        
        material_aggregate_df.sort_values(['system_code'], inplace = True)
        material_aggregate_df.drop(['year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december'], inplace=True, axis = 1)

        labour_aggregate_df.sort_values(['system_code'], inplace = True)
        labour_aggregate_df.drop(['year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december'], inplace=True, axis = 1)

        purchase_aggregate_df.sort_values(['system_code'], inplace =True)

        completion_rate_df.sort_values(['system_code'], inplace =True)
        master_aggregate.sort_values(['system_code'], inplace =True)

        required_df = pd.merge(labour_aggregate_df,material_aggregate_df, how = 'outer', on = ['system_code'] ).merge(purchase_aggregate_df, how = 'outer', on = ['system_code'] ).merge( master_aggregate,  how = 'outer', on = ['system_code'] )
        required_df.drop(['id_x', 'id_y', 'id'], axis = 1, inplace = True)
        required_df.fillna(0, inplace=True)
        required_df.rename(columns={'total_x': 'Labour_Cost', 'total_y': 'Material_Cost', 'total':'Purchase_Orders', 'master_total':'total_NRC' }, inplace = True)  
        required_df = required_df.merge(completion_rate_df, how ='left', on =['system_code'])   
        required_df.drop(['id'], axis = 1, inplace =True)   
        required_df.fillna(0, inplace =True) #The required table to be displayed
        #print(required_df.tail(50))

        #Migrating data from master_yearly_total_df to master_yearly_total table
        model= masterYearlyTotal.objects.all()
        model.delete()
        for index, row in master_yearly_total_df.iterrows():
               model = masterYearlyTotal()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january = round(row['january'],2)
               model.february = round(row['february'],2)
               model.march = round(row['march'],2)
               model.april = round(row['april'],2)
               model.may = round(row['may'],2)
               model.june = round(row['june'],2)
               model.july = round(row['july'],2)
               model.august = round(row['august'],2)
               model.september = round(row['september'],2)
               model.october = round(row['october'],2)
               model.november = round(row['november'],2)
               model.december = round(row['december'],2)
               model.master_total = round(row['master_total'],2)
               model.save()

        #Migrating data from master_aggregate data frame to master_aggregate table
        model= masterAggregate.objects.all()
        model.delete()
        for index, row in master_aggregate.iterrows():
               model = masterAggregate()
               model.system_code = row['system_code']
               model.master_total = round(row['master_total'],2)
               model.save()

        #Migrating data from master_aggregate data frame to required table
        model= requiredTable.objects.all()
        model.delete()
        for index, row in required_df.iterrows():
               model = requiredTable()
               model.system_code = row['system_code']
               model.labour_cost = round( row['Labour_Cost'], 2)
               model.material_cost = round(row['Material_Cost'],2)
               model.purchase_orders = round(row['Purchase_Orders'], 2)
               model.total_nrc = round(row['total_NRC'],2)
               model.completion_rate = round(float(row['completion_rate']),2) if row['completion_rate'] != '' and row['completion_rate'] != 'not flight relevant' else 0
               model.save()

        
        return render(request, 'index.html')
          
      else:
          return render(request, 'index.html')

    else:
 
       return render(request, 'index.html')

def dl_operations(request):
  #Reserved for A.I. Operations
  if(request.method == 'POST'):
      accu_tol = pd.DataFrame(columns = ['accuracy', 'tolerance'])
      if(request.POST.get('AI')):
        material_master_two = materialMaster.objects.all().values()
        labour_master_two = labourMaster.objects.all().values()
        ai_one = AITableOne.objects.all().values()
        completion = requiredTable.objects.all().values()
        total = masterYearlyTotal.objects.all().values()

        material_master_two_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        labour_master_two_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        ai_one_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
       
        ai_trainer_data = pd.DataFrame()
        material_master_two_df = pd.DataFrame(material_master_two)
        labour_master_two_df = pd.DataFrame( labour_master_two)
        ai_one_df = pd.DataFrame(ai_one)
        completion_df = pd.DataFrame(completion)
        completion_df = completion_df[['system_code', 'completion_rate']]
        total_df = pd.DataFrame(total)
        labour_master_two_df_copy = pd.DataFrame()
        material_master_two_df_copy = pd.DataFrame()
        total_df_copy = pd.DataFrame()
        ai_one_df_copy = pd.DataFrame()
        dummy_df = pd.DataFrame()
        ai_one_df_update = pd.DataFrame()
        ai_trainer_data_copy = pd.DataFrame()
        features = pd.DataFrame()
        total_appender = pd.DataFrame(columns = ['system_code', 'year', 'master_total'])
        ai_one_dummy_df = pd.DataFrame()
        ai_one_appender = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        labour_master_two_df_update = pd.DataFrame()
        total_df_update = pd.DataFrame()
        total_dummy_df = pd.DataFrame()
        labour_master_two_dummy_df = pd.DataFrame()
        labour_master_two_appender = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
        material_master_two_df_update = pd.DataFrame()
        material_master_two_dummy_df = pd.DataFrame()
        material_master_two_appender = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
       
        #Preparing cumulative moving average data for the ai_one data frame
        
        sys_code_list = ai_one_df['system_code'].unique()

        ai_one_df_update = ai_one_df.groupby(['system_code', 'year'], as_index = False).sum()
        ai_one_df_update = ai_one_df_update[['system_code', 'year']]
        current_year = date.today().year
        for sys_code in sys_code_list:
          ai_one_dummy_df = ai_one_df_update.loc[ai_one_df_update.loc[:,'system_code'] == sys_code ]
          if current_year in ai_one_dummy_df['year'].values :
            continue
          else:
            ai_one_appender.at[0, 'system_code'] = sys_code
            ai_one_appender.at[0, 'year'] = current_year
            ai_one_appender.at[0, 'january'] = 0
            ai_one_appender.at[0, 'february'] = 0
            ai_one_appender.at[0, 'march'] = 0
            ai_one_appender.at[0, 'april'] = 0
            ai_one_appender.at[0, 'may'] = 0
            ai_one_appender.at[0, 'june'] = 0
            ai_one_appender.at[0, 'july'] = 0
            ai_one_appender.at[0, 'august'] = 0
            ai_one_appender.at[0, 'september'] = 0
            ai_one_appender.at[0, 'october'] = 0
            ai_one_appender.at[0, 'november'] = 0
            ai_one_appender.at[0, 'december'] = 0
            ai_one_appender.at[0, 'master_total'] = 0
          
            ai_one_df = ai_one_df.append(ai_one_appender)
            ai_one_appender = ai_one_appender[0:0]
            ai_one_dummy_df = ai_one_dummy_df[0:0]

        
        ai_one_df.reset_index(inplace = True, drop = True) 

        print(ai_one_df.tail(50))
         
        for sys_code in sys_code_list:
        
          dummy_df = ai_one_df.loc[ai_one_df.loc[:,'system_code'] == sys_code ]
          dummy_df.reset_index(drop = True, inplace = True)
          dummy_df = dummy_df.sort_values(['year'])
          dummy_df['january']= dummy_df['january'].expanding().mean()
          dummy_df['february']= dummy_df['february'].expanding().mean()
          dummy_df['march']= dummy_df['march'].expanding().mean()
          dummy_df['april']= dummy_df['april'].expanding().mean()
          dummy_df['may']= dummy_df['may'].expanding().mean()
          dummy_df['june']= dummy_df['june'].expanding().mean()
          dummy_df['july']= dummy_df['july'].expanding().mean()
          dummy_df['august']= dummy_df['august'].expanding().mean()
          dummy_df['september']= dummy_df['september'].expanding().mean()
          dummy_df['october']= dummy_df['october'].expanding().mean()
          dummy_df['november']= dummy_df['november'].expanding().mean()
          dummy_df['december']= dummy_df['december'].expanding().mean()
          ai_one_df_copy = ai_one_df_copy.append(dummy_df)
          dummy_df = dummy_df[0:0]
        dummy_df = dummy_df[0:0]
        ai_one_df_copy.reset_index(drop =True, inplace =True)
        ai_one_df_copy.drop(['id', 'master_total'], axis = 1, inplace =True)
       

        
        #Preparing cumulative moving average data for the labour master data frame
        labour_master_two_df.drop(['id', 'total'], axis = 1, inplace =True)
        
        sys_code_list = labour_master_two_df['system_code'].unique()

        labour_master_two_df_update = labour_master_two_df.groupby(['system_code', 'year'], as_index = False).sum()
        labour_master_two_df_update = labour_master_two_df_update[['system_code', 'year']]
        current_year = date.today().year
        for sys_code in sys_code_list:
          labour_master_two_dummy_df = labour_master_two_df_update.loc[labour_master_two_df_update.loc[:,'system_code'] == sys_code ]
          if current_year in labour_master_two_dummy_df['year'].values :
            continue
          else:
            labour_master_two_appender.at[0, 'system_code'] = sys_code
            labour_master_two_appender.at[0, 'year'] = current_year
            labour_master_two_appender.at[0, 'january'] = 0
            labour_master_two_appender.at[0, 'february'] = 0
            labour_master_two_appender.at[0, 'march'] = 0
            labour_master_two_appender.at[0, 'april'] = 0
            labour_master_two_appender.at[0, 'may'] = 0
            labour_master_two_appender.at[0, 'june'] = 0
            labour_master_two_appender.at[0, 'july'] = 0
            labour_master_two_appender.at[0, 'august'] = 0
            labour_master_two_appender.at[0, 'september'] = 0
            labour_master_two_appender.at[0, 'october'] = 0
            labour_master_two_appender.at[0, 'november'] = 0
            labour_master_two_appender.at[0, 'december'] = 0
            labour_master_two_appender.at[0, 'total'] = 0
          
            labour_master_two_df = labour_master_two_df.append(labour_master_two_appender)
            labour_master_two_appender = labour_master_two_appender[0:0]
            labour_master_two_dummy_df = labour_master_two_dummy_df[0:0]

        
        
        labour_master_two_df.reset_index(inplace = True, drop = True) 
    
        for sys_code in sys_code_list:
          dummy_df = labour_master_two_df.loc[labour_master_two_df.loc[:,'system_code'] == sys_code ]
          dummy_df.reset_index(drop = True, inplace = True)
          dummy_df = dummy_df.sort_values(['year'])
          dummy_df['january']= dummy_df['january'].expanding().mean()
          dummy_df['february']= dummy_df['february'].expanding().mean()
          dummy_df['march']= dummy_df['march'].expanding().mean()
          dummy_df['april']= dummy_df['april'].expanding().mean()
          dummy_df['may']= dummy_df['may'].expanding().mean()
          dummy_df['june']= dummy_df['june'].expanding().mean()
          dummy_df['july']= dummy_df['july'].expanding().mean()
          dummy_df['august']= dummy_df['august'].expanding().mean()
          dummy_df['september']= dummy_df['september'].expanding().mean()
          dummy_df['october']= dummy_df['october'].expanding().mean()
          dummy_df['november']= dummy_df['november'].expanding().mean()
          dummy_df['december']= dummy_df['december'].expanding().mean()
          labour_master_two_df_copy = labour_master_two_df_copy.append(dummy_df)
          dummy_df = dummy_df[0:0]
        dummy_df = dummy_df[0:0]
        labour_master_two_df_copy.reset_index(drop =True, inplace =True)


        #Preparing cumulative moving average data for the material master data frame
        material_master_two_df.drop(['id', 'total'], axis = 1, inplace = True)
        
        sys_code_list = material_master_two_df['system_code'].unique()
        
        material_master_two_df_update = material_master_two_df.groupby(['system_code', 'year'], as_index = False).sum()
        material_master_two_df_update = material_master_two_df_update[['system_code', 'year']]
        current_year = date.today().year
        for sys_code in sys_code_list:
          material_master_two_dummy_df = material_master_two_df_update.loc[material_master_two_df_update.loc[:,'system_code'] == sys_code ]
          if current_year in material_master_two_dummy_df['year'].values :
            continue
          else:
            material_master_two_appender.at[0, 'system_code'] = sys_code
            material_master_two_appender.at[0, 'year'] = current_year
            material_master_two_appender.at[0, 'january'] = 0
            material_master_two_appender.at[0, 'february'] = 0
            material_master_two_appender.at[0, 'march'] = 0
            material_master_two_appender.at[0, 'april'] = 0
            material_master_two_appender.at[0, 'may'] = 0
            material_master_two_appender.at[0, 'june'] = 0
            material_master_two_appender.at[0, 'july'] = 0
            material_master_two_appender.at[0, 'august'] = 0
            material_master_two_appender.at[0, 'september'] = 0
            material_master_two_appender.at[0, 'october'] = 0
            material_master_two_appender.at[0, 'november'] = 0
            material_master_two_appender.at[0, 'december'] = 0
            material_master_two_appender.at[0, 'total'] = 0
          
            material_master_two_df = material_master_two_df.append(material_master_two_appender)
            material_master_two_appender = material_master_two_appender[0:0]
            material_master_two_dummy_df = material_master_two_dummy_df[0:0]

        
        
        material_master_two_df.reset_index(inplace = True, drop = True) 



        for sys_code in sys_code_list:
            dummy_df = material_master_two_df.loc[material_master_two_df.loc[:,'system_code'] == sys_code ]
            dummy_df.reset_index(drop = True, inplace = True)
            dummy_df = dummy_df.sort_values(['year'])
            dummy_df['january']= dummy_df['january'].expanding().mean()
            dummy_df['february']= dummy_df['february'].expanding().mean()
            dummy_df['march']= dummy_df['march'].expanding().mean()
            dummy_df['april']= dummy_df['april'].expanding().mean()
            dummy_df['may']= dummy_df['may'].expanding().mean()
            dummy_df['june']= dummy_df['june'].expanding().mean()
            dummy_df['july']= dummy_df['july'].expanding().mean()
            dummy_df['august']= dummy_df['august'].expanding().mean()
            dummy_df['september']= dummy_df['september'].expanding().mean()
            dummy_df['october']= dummy_df['october'].expanding().mean()
            dummy_df['november']= dummy_df['november'].expanding().mean()
            dummy_df['december']= dummy_df['december'].expanding().mean()
            material_master_two_df_copy = material_master_two_df_copy.append(dummy_df)
            dummy_df = dummy_df[0:0]
        dummy_df = dummy_df[0:0] 
        material_master_two_df_copy.reset_index(drop =True, inplace =True)

        #Fusing purchase_costs with material and labour costs, based on purchase cost
      
        ai_trainer_data = pd.merge(ai_one_df_copy, labour_master_two_df_copy , how = 'outer', on = ['system_code', 'year'] )
        ai_trainer_data = ai_trainer_data.merge(material_master_two_df_copy, how = 'outer', on = ['system_code', 'year'] )
        ai_trainer_data = ai_trainer_data.merge(completion_df, how = 'left', on = ['system_code'])
       
        total_df = total_df[['system_code','year', 'master_total']]

        sys_code_list = total_df['system_code'].unique()


        total_df_update = total_df.groupby(['system_code', 'year'], as_index = False).sum()
        total_df_update = total_df_update[['system_code', 'year']]
        current_year = date.today().year
        for sys_code in sys_code_list:
          total_dummy_df = total_df_update.loc[total_df_update.loc[:,'system_code'] == sys_code ]
          if current_year in total_dummy_df['year'].values :
            continue
          else:
            total_appender.at[0, 'system_code'] = sys_code
            total_appender.at[0, 'year'] = current_year
            total_appender.at[0, 'master_total'] = 0
            total_df = total_df.append(total_appender)
            total_appender = total_appender[0:0]
            total_dummy_df = total_dummy_df[0:0]

        
        total_df.reset_index(inplace = True, drop = True) 


        for sys_code in sys_code_list:
            dummy_df = total_df.loc[total_df.loc[:,'system_code'] == sys_code ]
            dummy_df.reset_index(drop = True, inplace = True)
            dummy_df = dummy_df.sort_values(['year'])
            dummy_df['master_total']= dummy_df['master_total'].expanding().mean()
            total_df_copy = total_df_copy.append(dummy_df)
            dummy_df = dummy_df[0:0]
        dummy_df = dummy_df[0:0] 
        total_df_copy.reset_index(drop =True, inplace =True)


        ai_trainer_data = ai_trainer_data.merge(total_df_copy, how = 'left', on = ['system_code', 'year'])
        ai_trainer_data.fillna(0, inplace = True)

        #Migrating the data from ai_trainer_data frame to ai_trainer_data table in the database
        model= AITrainerData.objects.all()
        model.delete()
        for index, row in ai_trainer_data.iterrows():
             
               model = AITrainerData()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january_po = row['january_x']
               model.february_po = row['february_x']
               model.march_po = row['march_x']
               model.april_po = row['april_x']
               model.may_po = row['may_x']
               model.june_po = row['june_x']
               model.july_po = row['july_x']
               model.august_po = row['august_x']
               model.september_po = row['september_x']
               model.october_po = row['october_x']
               model.november_po = row['november_x']
               model.december_po = row['december_x']
               model.january_lb = row['january_y']
               model.february_lb = row['february_y']
               model.march_lb = row['march_y']
               model.april_lb = row['april_y']
               model.may_lb = row['may_y']
               model.june_lb = row['june_y']
               model.july_lb = row['july_y']
               model.august_lb = row['august_y']
               model.september_lb = row['september_y']
               model.october_lb = row['october_y']
               model.november_lb = row['november_y']
               model.december_lb = row['december_y']
               model.january_mt = row['january']
               model.february_mt = row['february']
               model.march_mt = row['march']
               model.april_mt = row['april']
               model.may_mt = row['may']
               model.june_mt = row['june']
               model.july_mt = row['july']
               model.august_mt = row['august']
               model.september_mt = row['september']
               model.october_mt = row['october']
               model.november_mt = row['november']
               model.december_mt = row['december']
               model.completion_rate = row['completion_rate']
               model.master_total = row['master_total']
               model.save()

        print("AI traine data")       
        print(" ")
        print(ai_trainer_data.head(50))

        #Normalizing the A.I. Trainer Data Frame
        ai_trainer_data_copy = ai_trainer_data.copy()
        print(ai_trainer_data_copy.head(50))
        ai_trainer_data_copy[['january_x', 'february_x', 'march_x', 'april_x', 'may_x', 'june_x', 'july_x', 'august_x', 'september_x', 'october_x', 'november_x', 'december_x', 'january_y', 'february_y', 'march_y', 'april_y', 'may_y', 'june_y', 'july_y', 'august_y', 'september_y', 'october_y', 'november_y', 'december_y', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december' ]]= MinMaxScaler().fit_transform(ai_trainer_data_copy[['january_x', 'february_x', 'march_x', 'april_x', 'may_x', 'june_x', 'july_x', 'august_x', 'september_x', 'october_x', 'november_x', 'december_x', 'january_y', 'february_y', 'march_y', 'april_y', 'may_y', 'june_y', 'july_y', 'august_y', 'september_y', 'october_y', 'november_y', 'december_y', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december' ]])
      

        #Migrating the data from ai_trainer_data_normalized frame to ai_trainer_data_normalized table in the database
        model= AITrainerDataScaled.objects.all()
        model.delete()
        for index, row in ai_trainer_data_copy.iterrows():
               model = AITrainerDataScaled()
               model.system_code = row['system_code']
               model.year = row['year']
               model.january_po = row['january_x']
               model.february_po = row['february_x']
               model.march_po = row['march_x']
               model.april_po = row['april_x']
               model.may_po = row['may_x']
               model.june_po = row['june_x']
               model.july_po = row['july_x']
               model.august_po = row['august_x']
               model.september_po = row['september_x']
               model.october_po = row['october_x']
               model.november_po = row['november_x']
               model.december_po = row['december_x']
               model.january_lb = row['january_y']
               model.february_lb = row['february_y']
               model.march_lb = row['march_y']
               model.april_lb = row['april_y']
               model.may_lb = row['may_y']
               model.june_lb = row['june_y']
               model.july_lb = row['july_y']
               model.august_lb = row['august_y']
               model.september_lb = row['september_y']
               model.october_lb = row['october_y']
               model.november_lb = row['november_y']
               model.december_lb = row['december_y']
               model.january_mt = row['january']
               model.february_mt = row['february']
               model.march_mt = row['march']
               model.april_mt = row['april']
               model.may_mt = row['may']
               model.june_mt = row['june']
               model.july_mt = row['july']
               model.august_mt = row['august']
               model.september_mt = row['september']
               model.october_mt = row['october']
               model.november_mt = row['november']
               model.december_mt = row['december']
               model.completion_rate = row['completion_rate']
               model.master_total = row['master_total']
               model.save()


        print('\n')
        print(ai_trainer_data_copy.head(50))

        return render(request, 'dl_operations.html')
      
      if request.POST.get('NN'):
        np.set_printoptions(suppress = True)
        accuracy = 'Training...'
        current_year = date.today().year
        trainer_data = AITrainerData.objects.all().values()
        data = pd.DataFrame(trainer_data)
        
        
        data_train = data.loc[~(data['year'] == current_year)]
        labels = pd.DataFrame()
        

        multiplier = int(request.POST.get('mul'))
        data_train = pd.concat([data_train]* multiplier, ignore_index= True)
        data_train = data_train.sample(frac = 1)



        labels['master_total'] =data_train.loc[:, 'master_total']

        features = data_train.iloc[:,0:40]

        le=preprocessing.LabelEncoder()
        features['system_code'] = le.fit_transform(features['system_code'])
        features = features.drop(['year'], axis = 1)
        features = features.drop(['id'], axis = 1)

        print(features.head(50))
        
        features.loc[:,'january_po': 'december_mt'] = MinMaxScaler().fit_transform(features.loc[:,'january_po': 'december_mt'])
        features.reset_index(drop = True, inplace = True)
        labels.reset_index(drop = True, inplace = True)

        

        X = features.loc[:,'system_code':'completion_rate'].values
        y = labels['master_total'].values
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.33, random_state=42)



        model = Sequential()
        model.add(Dense(256, activation = 'relu', input_dim = 38))
        model.add(Dense(128, activation ='relu'))
        
        model.add(Dense(32, activation = 'relu'))
        model.add(Dense(8, activation = 'relu'))
        model.add(Dense (1, activation = "linear"))

        #model training
        epochs = int(request.POST.get('epochs'))
        model.compile( loss = 'mean_squared_error', optimizer= adam_v2.Adam(learning_rate = 0.001) , metrics = ['accuracy'])
        model.fit(X_train, y_train, epochs = epochs, batch_size = 1, verbose =1)
        model.save('my_model.h5')

        y_pred = model.predict(X_test)
        print(y_test[0:50])
        print(y_pred[0:50])
        tolerance = int(request.POST.get('tolerance'))
        accuracy = (np.abs(y_test - abs(y_pred))  < tolerance).mean()
        score = model.evaluate(X_test, y_test,verbose=1)
      
        accuracy = round(accuracy,2) * 100 

        accu_tol = accu_tol[0:0]
        accu_tol.at[0,'accuracy'] = accuracy
        accu_tol.at[0,'tolerance'] = tolerance
        
        model= accuracyTolerance.objects.all()
        model.delete()
        for index, row in accu_tol.iterrows():
               model = accuracyTolerance()
               model.accuracy = accu_tol['accuracy']
               model.tolerance = accu_tol['tolerance']
               
               model.save()
       
        print(accu_tol)


      
        return render(request, 'dl_operations.html', context = {'accuracy':accuracy, 'toler': tolerance})

      if(request.POST.get('syscode') and request.POST.get('predict')):

        syscode = request.POST.get('syscode')
        pred = pd.DataFrame()
        
        store = accuracyTolerance.objects.all().values()
        accuracyTot = pd.DataFrame(store)

        np.set_printoptions(suppress = True)
       
        current_year = date.today().year
        prediction_data = AITrainerData.objects.all().values()
        labels_predict = pd.DataFrame()
        data = pd.DataFrame(prediction_data)

        data_predict = data.loc[data['year']== current_year]
        labels_predict['master_total'] = data_predict.loc[:, 'master_total']

        features_predict = data_predict.iloc[:,0:40]

        le=preprocessing.LabelEncoder()
        features_predict['system_code'] = le.fit_transform(features_predict['system_code'])
        features_predict = features_predict.drop(['year'], axis = 1)

        features_predict.loc[:,'january_po': 'december_mt'] = MinMaxScaler().fit_transform(features_predict.loc[:,'january_po': 'december_mt'])
        features_predict.reset_index(drop = True, inplace = True)
        labels_predict.reset_index(drop = True, inplace = True)

        X_predict = features_predict.loc[:,'system_code':'completion_rate'].values
        y_predict = labels_predict['master_total'].values

        my_model = load_model('my_model.h5')

        y_predict_ai = my_model.predict(X_predict)
        print(accuracyTot)
        list_accuTot = accuracyTot.values.tolist()
        list_accuTot = np.array(list_accuTot)
        list_accuTot = list_accuTot.flatten()


        print(list_accuTot)
        accuracy = list_accuTot[1]
        
        tolerance = int(list_accuTot[2])
        
        error = (100 - accuracy)/100
        features_predict['system_code'] = le.inverse_transform(features_predict['system_code'])
        features_predict = features_predict['system_code']
        predictions =pd.DataFrame(y_predict_ai, columns = ['predictions'])
        final_df = pd.concat([features_predict, predictions], axis = 1)

        print(final_df.head(50))
        
        systemcodes = final_df['system_code'].unique()

        if( syscode in systemcodes):
          pred = final_df.loc[final_df['system_code']==syscode]
          pred_value= str(pred.loc[:,'predictions'])
          pred_value = pred_value.split()
          pred_value = pred_value[1].rstrip('.')
          pred_value = round(float(pred_value), 2)
          lower = round((pred_value - (error * pred_value)), 2)
          upper = round((pred_value + (error * pred_value)), 2)
          print(pred_value)
       
          return render(request, 'dl_operations.html', context = {'predictions': pred_value, 'syscode':syscode, 'year': current_year, 'lower': lower, 'upper': upper, 'tolerance': tolerance, 'accuracy':accuracy  })
        else:
          return render(request, 'dl_operations.html', context = {'predictions':'System Code Not Found', 'syscode':syscode, 'year': current_year, 'accuracy':accuracy  })

      if (request.POST.get('forecast') and request.POST.get('tolerance2') and request.POST.get('duration') and request.POST.get('seasonal') ):
          
          data = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        
          store = masterYearlyTotal.objects.all().values()
          data = pd.DataFrame(store)
      
          data.drop(['system_code', 'master_total'], axis = 1, inplace = True)
          labels = {'january':'sum', 'february': 'sum', 'march': 'sum', 'april': 'sum', 'may': 'sum', 'june':'sum', 'july':'sum', 'august':'sum', 'september':'sum', 'october':'sum', 'november':'sum', 'december': 'sum'}
          data = data.groupby(['year'], as_index= False).agg(labels)
          years = data['year'].unique()
          data_narrowed =pd.DataFrame()
          data_transposed = data.T
          data_transposed = data_transposed.iloc[1:,:]
       

          n_columns = data_transposed.shape[1]

          col = data_transposed.iloc[:, 3]

          for cols in range(0,(n_columns)):
              col = data_transposed.iloc[:, cols]
              col = pd.DataFrame(col)
              col.reset_index(level = 0, inplace = True)
            
          
              data_narrowed = data_narrowed.append(col)   

          data_narrowed.reset_index(drop = True, inplace =True)
          data_narrowed = data_narrowed.fillna(0)
          data_narrowed['expenditure'] = data_narrowed.iloc[:, 0:].sum(axis = 1)
          data_narrowed = data_narrowed[['index', 'expenditure']]

          years = np.repeat(years,12)
          years = pd.DataFrame({'years': years})
          data_narrowed = pd.concat([data_narrowed, years], axis = 1)
          data_narrowed['index' ] = data_narrowed['index'].str[:3]
          data_narrowed['date'] = data_narrowed['index'].str.capitalize() + "-" + data_narrowed['years'].astype(str)

          data_narrowed['date'] = pd.to_datetime(data_narrowed['date'])
          data_narrowed = data_narrowed[['date', 'expenditure']]
          data_narrowed = data_narrowed[data_narrowed['expenditure'] != 0]


          data_narrowed.set_index('date',inplace=True)
          data_narrowed['expenditure_first_difference'] = data_narrowed['expenditure'] - data_narrowed['expenditure'].shift(3)
          data_narrowed = data_narrowed.dropna()

          model2=sm.tsa.statespace.SARIMAX(data_narrowed['expenditure'],order=(1, 1, 1),seasonal_order=(1,1,1,int(request.POST.get('seasonal'))))
          results=model2.fit()

          data_narrowed['forecast']= results.predict(start=15,end=40,dynamic=True)
          
          future_dates=[data_narrowed.index[-1] + DateOffset(months=x) for x in range(0, int(request.POST.get('duration')))]



          future_dataset_df = pd.DataFrame(index=future_dates[1:], columns = data_narrowed.columns)

          future_df = pd.concat([data_narrowed, future_dataset_df])
          maximum = len(data_narrowed)

          future_df['forecast']= results.predict(start= int(request.POST.get('initializer')),end= len(data_narrowed) + int(request.POST.get('duration')),dynamic=True)
          future_df[['expenditure','forecast']].plot(figsize=(12,8))
          pyplot.xlabel("Years")
          pyplot.ylabel("Expenditure in euros")
          pyplot.title("Forecast Fitting")
          pyplot.show()
          initializer = int(request.POST.get('initializer'))

          expenditure_df = pd.DataFrame(columns = future_df.columns)
          future_forecast_df = pd.DataFrame(columns = future_df.columns)
          current_date = date.today()
          future_df = future_df.fillna(0)

          expenditure_df = future_df.loc[future_df['expenditure'] != 0]
          future_forecast_df = future_df.loc[future_df['expenditure'] == 0]


          real_sum = expenditure_df['expenditure'].sum()
          forecast_sum = future_forecast_df['forecast'].sum()

          total_expenditure = round((real_sum + forecast_sum),2)

         
          data_narrowed = data_narrowed.fillna(0)
          util_df = data_narrowed.loc[data_narrowed['forecast'] != 0]
          exp_actual_list = util_df['expenditure'].values


          exp_predicted_list = util_df['forecast'].values

          util_df.reset_index( drop = False, inplace = True)

          future_df = future_df.loc[future_df['expenditure'] == 0]
          future_df.reset_index(drop = False, inplace = True)
          future_df['index'] = future_df['index'].astype(str)
          print(future_df)

          selected_period = str(future_df['index'].to_list()[-1]) if len(future_df['index']) > 0 else "Invalid Period"

          json_records = future_df.reset_index().to_json(orient = 'records')
          json_data = []
          json_data = json.loads(json_records)
         
          

          tolerance_predict = int(request.POST.get('tolerance2'))
          
          accuracy_predict =(round((np.abs( exp_predicted_list[:17] - abs(exp_actual_list[:17]))  < tolerance_predict).mean(), 2) *100)

          error_predict =  (100 - accuracy_predict)/100

          lower_predict = round((total_expenditure - (error_predict * total_expenditure)), 2)
          upper_predict = round((total_expenditure + (error_predict * total_expenditure)), 2)
          
          return render(request, 'dl_operations.html', context = { 'initializer': initializer, 'seasonal':int(request.POST.get('seasonal')) ,'max': maximum, 'remnant_expenditure' : round(forecast_sum,2), 'pred_data': json_data,  'tolerance2': tolerance_predict, 'accuracy2':accuracy_predict,  'total_expenditure' : total_expenditure, 'lower_predict': lower_predict, 'upper_predict' : upper_predict, 'selected_period': selected_period})

      if (request.POST.get('time_steps') and request.POST.get('epoch2') and request.POST.get('train_test') and request.POST.get('tolerance3')  and request.POST.get('forecast2') ):

  
          data = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        
          store = masterYearlyTotal.objects.all().values()
          data = pd.DataFrame(store)
        
          data.drop(['system_code', 'master_total'], axis = 1, inplace = True)
          labels = {'january':'sum', 'february': 'sum', 'march': 'sum', 'april': 'sum', 'may': 'sum', 'june':'sum', 'july':'sum', 'august':'sum', 'september':'sum', 'october':'sum', 'november':'sum', 'december': 'sum'}
          data = data.groupby(['year'], as_index= False).agg(labels)
          years = data['year'].unique()
          data_narrowed =pd.DataFrame()
          data_transposed = data.T
          data_transposed = data_transposed.iloc[1:,:]
          print(data.head())
          print(data_transposed)


          n_columns = data_transposed.shape[1]

          col = data_transposed.iloc[:, 3]

          for cols in range(0,(n_columns)):
              col = data_transposed.iloc[:, cols]
              col = pd.DataFrame(col)
              col.reset_index(level = 0, inplace = True)
            
          
              data_narrowed = data_narrowed.append(col)   

          data_narrowed.reset_index(drop = True, inplace =True)
          data_narrowed = data_narrowed.fillna(0)
          data_narrowed['expenditure'] = data_narrowed.iloc[:, 0:].sum(axis = 1)
          data_narrowed = data_narrowed[['index', 'expenditure']]

          years = np.repeat(years,12)
          years = pd.DataFrame({'years': years})
          data_narrowed = pd.concat([data_narrowed, years], axis = 1)
          data_narrowed['index' ] = data_narrowed['index'].str[:3]
          data_narrowed['date'] = data_narrowed['index'].str.capitalize() + "-" + data_narrowed['years'].astype(str)
          data_narrowed['date'] = pd.to_datetime(data_narrowed['date'])
          data_narrowed = data_narrowed[['date', 'expenditure']]
          data_narrowed = data_narrowed[data_narrowed['expenditure'] != 0]

          data_narrowed.set_index('date',inplace=True)

          train_size = int(len(data_narrowed) * float(request.POST.get('train_test')))
          test_size = len(data_narrowed) - train_size
          train, test = data_narrowed.iloc[0:train_size], data_narrowed.iloc[train_size : len(data_narrowed)]
         
          transformer = RobustScaler()
          transformer = transformer.fit(train[['expenditure']])
          train['expenditure'] = transformer.transform(train[['expenditure']])
          test['expenditure'] = transformer.transform(test[['expenditure']])

          def create_dataset(X, y, time_steps=1):
              Xs, ys = [], []
              for i in range(len(X) - time_steps):
                  v = X.iloc[i:(i + time_steps)].values
                  Xs.append(v)
                  ys.append(y.iloc[i + time_steps])
              return np.array(Xs), np.array(ys)

          time_steps =  int(request.POST.get('time_steps'))

          # reshape to [samples, time_steps, n_features]

          X_train, y_train = create_dataset(train, train.expenditure, time_steps)
          X_test, y_test = create_dataset(test, test.expenditure, time_steps)

          model =keras.Sequential()
          model.add(
            keras.layers.Bidirectional(
              keras.layers.LSTM(
                units = 256,
                input_shape = (X_train.shape[1], X_train.shape[2]))))
          model.add(keras.layers.Dense(units=128))
          model.add(keras.layers.Dense(units=1))
          model.compile(loss='mean_squared_error', optimizer= adam_v2.Adam(0.001))

          history = model.fit(X_train, y_train, epochs= int(request.POST.get('epoch2')), batch_size=1, validation_split=0.1, shuffle=False)
          plt.plot(history.history['loss'], label = 'train')
          plt.plot(history.history['val_loss'], label = 'test')

          plt.xlabel("Epochs")
          plt.ylabel("Loss")
          plt.title("Loss VS Epochs for training (Blue) and Validation (Orange) Data")
          plt.show()
          y_pred = model.predict(X_test)

          y_train_inv = transformer.inverse_transform(y_train.reshape(-1,1))
          y_test_inv = transformer.inverse_transform(y_test.reshape(-1,1))
          y_pred_inv = transformer.inverse_transform(y_test.reshape(-1,1))

          future_pred = data_narrowed.iloc[-time_steps: ]
          future_pred_list = future_pred.values
          future_pred_list = future_pred_list.reshape(1,time_steps,1)

          future_prediction = model.predict(future_pred_list)
          #future_pred_list_inverse = transformer.inverse_transform(future_pred_list.reshape(1,-1))
          future_prediction_inverse = abs(transformer.inverse_transform(future_prediction))
          predicted_cost = future_prediction_inverse.flatten()[0]
          predicted_cost = round(predicted_cost,2)
          future_date=[data_narrowed.index[-1] + DateOffset(months=1)]
          tolerance_predict = int(request.POST.get('tolerance3'))
          accuracy_predict = (np.abs( y_test_inv.flatten() - abs(y_pred_inv.flatten()))  < tolerance_predict).mean() * 100


          error_predict =  (100 - accuracy_predict)/100

          lower_predict = (predicted_cost - (error_predict * predicted_cost))
          upper_predict = (predicted_cost + (error_predict * predicted_cost))


          
          return render(request, 'dl_operations.html', context = { 'accuracy3': accuracy_predict, 'time': str(future_date[0]).split(' ')[0], 'tolerance3': tolerance_predict,  'lower_predict2': lower_predict, 'upper_predict2' : upper_predict, 'total_expenditure2': predicted_cost, 'epoch2': request.POST.get('epoch2') })

      if( request.POST.get('generate') and request.POST.get('seasonal2') and request.POST.get('initializer2') and request.POST.get('duration2')):
     
        current_year = date.today().year
        seasonal2 = int(request.POST.get('seasonal2'))
        initializer2 = int(request.POST.get('initializer2'))
        duration2 = int(request.POST.get('duration2'))

        data = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        data2 = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        values = masterYearlyTotal.objects.all().values()
        data = pd.DataFrame(values)
        data2 = pd.DataFrame(values)
        data.drop(['id'], axis = 1, inplace = True)
        data2.drop(['id'], axis = 1, inplace = True )
        

        data.drop(['system_code', 'master_total'], axis = 1, inplace = True)
        labels = {'january':'sum', 'february': 'sum', 'march': 'sum', 'april': 'sum', 'may': 'sum', 'june':'sum', 'july':'sum', 'august':'sum', 'september':'sum', 'october':'sum', 'november':'sum', 'december': 'sum'}
        data = data.groupby(['year'], as_index= False).agg(labels)
        years = data['year'].unique()
        data_narrowed =pd.DataFrame()
        data_transposed = data.T
        data_transposed = data_transposed.iloc[1:,:]
    
    
        n_columns = data_transposed.shape[1]

        col = data_transposed.iloc[:, 3]

        for cols in range(0,(n_columns)):
            col = data_transposed.iloc[:, cols]
            col = pd.DataFrame(col)
            col.reset_index(level = 0, inplace = True)
          
        
            data_narrowed = data_narrowed.append(col)   

        data_narrowed.reset_index(drop = True, inplace =True)
        data_narrowed = data_narrowed.fillna(0)
        data_narrowed['expenditure'] = data_narrowed.iloc[:, 0:].sum(axis = 1)
        data_narrowed = data_narrowed[['index', 'expenditure']]

        years = np.repeat(years,12)
        years = pd.DataFrame({'years': years})
        data_narrowed = pd.concat([data_narrowed, years], axis = 1)
        data_narrowed['index' ] = data_narrowed['index'].str[:3]
        data_narrowed['date'] = data_narrowed['index'].str.capitalize() + "-" + data_narrowed['years'].astype(str)

        data_narrowed['date'] = pd.to_datetime(data_narrowed['date'])
        data_narrowed = data_narrowed[['date', 'expenditure']]
        data_narrowed = data_narrowed[data_narrowed['expenditure'] != 0]

        data_narrowed.set_index('date',inplace=True)
        
     

        syscode_list = data2['system_code'].unique()
        syscode_df = pd.DataFrame()
        for syscode in syscode_list:

              dummy_df = data2.loc[data2['system_code'] == syscode]
              dummy_df.drop(['system_code', 'master_total'], axis = 1, inplace = True)
              years = data['year'].unique()
              process_df =pd.DataFrame()
              transposed_df = dummy_df.T
              #print(transposed_df)
              transposed_df = transposed_df.iloc[1:,:]
              transposed_df.reset_index(drop = True, inplace = True)
              transposed_df.columns = range(transposed_df.columns.size)
              n_columns2 = transposed_df.shape[1]

              for cols in range(0,(n_columns2)):
                  col2 = transposed_df.iloc[:, cols]
                  col2 = pd.DataFrame(col2)
                  col2.reset_index(level = 0, inplace = True)
                  process_df = process_df.append(col2)  
              
              process_df.set_index('index', inplace = True)

              process_df.reset_index(drop = True, inplace =True)
              process_df = process_df.fillna(0)
              process_df[syscode] = process_df.iloc[:, 0:].sum(axis = 1)
              process_df = process_df[[ syscode]]
            
              syscode_df = pd.concat([syscode_df,process_df], axis = 1)
              process_df = process_df[0:0]

     
        data_narrowed.reset_index(drop=False, inplace = True)
        syscode_df = syscode_df[0:36]
        syscode_df = pd.concat([syscode_df, data_narrowed], axis = 1)
        syscode_df = syscode_df.fillna(0)
        syscode_df = syscode_df.set_index('date')


        syscode_df_copy = pd.DataFrame()
        syscode_df_copy = syscode_df.copy(deep = True)
        for syscode in syscode_list:
            syscode_df[syscode] =syscode_df[syscode].replace(0,syscode_df[syscode].mean() )
        forecast_df = pd.DataFrame()

        maxima = len(data_narrowed)
        future_dates=[syscode_df.index[-1] + DateOffset(months=x) for x in range(0,duration2)]

        for syscode in syscode_list:
                  model2=sm.tsa.statespace.SARIMAX(syscode_df[syscode],order=(1, 1, 1),seasonal_order=(1,1,1, seasonal2), enforce_stationarity=False)
                  results=model2.fit(method_kwargs={"warn_convergence": False})


                  future_dataset_df = pd.DataFrame(index=future_dates[1:], columns = syscode_df.columns)

                  future_df = pd.concat([syscode_df, future_dataset_df])

                  future_df[syscode+'_forecast']= results.predict(start=initializer2,end= maxima + int(request.POST.get('duration2')),dynamic=True)
                  future_df[future_df < 0] = 0
                  forecast_df = pd.concat([forecast_df,future_df.iloc[:,-1]], axis = 1)
                
                  future_df= future_df[0:0]
                  future_dataset_df = future_dataset_df[0:0]
        
        syscode_df = pd.concat([syscode_df,forecast_df], axis = 1)
        syscode_df.drop(['expenditure'], axis = 1, inplace = True)
        syscode_df[syscode_df < 0] = 0
        print(syscode_df.iloc[:,0:10])
        syscode_df.reset_index(drop=False, inplace = True)
        syscode_df['index'] = syscode_df['index'].astype(str)


        syscode_df_copy.drop([ 'expenditure'], axis = 1, inplace = True)


        columns = list(syscode_df_copy.columns)
        syscode_df_copy.reset_index(['date'], inplace = True)
        print(syscode_df_copy.head(50))

        syscode_df_copy2 = syscode_df

        syscode_df_copy2.drop([ syscode for syscode in columns ], inplace = True, axis = 1 )
        #syscode_df_copy2.drop(['system_code_first_difference'], inplace =True,axis = 1 )
        syscode_df_copy2.columns = syscode_df_copy2.columns.str.rstrip("_forecast")
        syscode_df_copy2.columns = syscode_df_copy2.columns.str.replace('Managemen', 'Management')
        syscode_df_copy2_new= syscode_df_copy2.rename(columns = {'index':'date'})
        print(syscode_df_copy2_new.tail(50))


        length= len(syscode_df_copy) -1
          

        syscode_df_copy2_new = syscode_df_copy2_new.iloc[length:,:]
        syscode_df_copy2_new.reset_index( drop = True, inplace = True)
        syscode_df_copy.loc[:,'Type'] = 'Actual'
        syscode_df_copy2_new.loc[:,'Type'] = 'SARIMAX'

        print(syscode_df_copy2_new.tail(50))
        syscode_df_copy['date'] = syscode_df_copy['date'].astype(str)
        syscode_df_copy = pd.concat([syscode_df_copy,syscode_df_copy2_new], axis = 0 )
        syscode_df_copy.reset_index(drop = True, inplace = True)
    
        
        print(syscode_df_copy.iloc[:,0:10])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="System_Code_level_Prediction.xlsx"'                                        
        syscode_df_copy.to_excel(response)
        return response


      if( request.POST.get('generate2') and request.POST.get('system_code00') and request.POST.get('seasonal2') and request.POST.get('initializer2') and request.POST.get('duration2') ):
        current_year = date.today().year
        seasonal2 = int(request.POST.get('seasonal2'))
        initializer2 = int(request.POST.get('initializer2'))
        duration2 = int(request.POST.get('duration2'))
        sysCode = request.POST.get('system_code00')
        data = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        data2 = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', ' june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
        values = masterYearlyTotal.objects.all().values()
        data = pd.DataFrame(values)
        data2 = pd.DataFrame(values)
        data.drop(['id'], axis = 1, inplace = True)
        data2.drop(['id'], axis = 1, inplace = True )
        

        data.drop(['system_code', 'master_total'], axis = 1, inplace = True)
        labels = {'january':'sum', 'february': 'sum', 'march': 'sum', 'april': 'sum', 'may': 'sum', 'june':'sum', 'july':'sum', 'august':'sum', 'september':'sum', 'october':'sum', 'november':'sum', 'december': 'sum'}
        data = data.groupby(['year'], as_index= False).agg(labels)
        years = data['year'].unique()
        data_narrowed =pd.DataFrame()
        data_transposed = data.T
        data_transposed = data_transposed.iloc[1:,:]
    
    
        n_columns = data_transposed.shape[1]

        col = data_transposed.iloc[:, 3]

        for cols in range(0,(n_columns)):
            col = data_transposed.iloc[:, cols]
            col = pd.DataFrame(col)
            col.reset_index(level = 0, inplace = True)
          
        
            data_narrowed = data_narrowed.append(col)   

        data_narrowed.reset_index(drop = True, inplace =True)
        data_narrowed = data_narrowed.fillna(0)
        data_narrowed['expenditure'] = data_narrowed.iloc[:, 0:].sum(axis = 1)
        data_narrowed = data_narrowed[['index', 'expenditure']]

        years = np.repeat(years,12)
        years = pd.DataFrame({'years': years})
        data_narrowed = pd.concat([data_narrowed, years], axis = 1)
        data_narrowed['index' ] = data_narrowed['index'].str[:3]
        data_narrowed['date'] = data_narrowed['index'].str.capitalize() + "-" + data_narrowed['years'].astype(str)

        data_narrowed['date'] = pd.to_datetime(data_narrowed['date'])
        data_narrowed = data_narrowed[['date', 'expenditure']]
        data_narrowed = data_narrowed[data_narrowed['expenditure'] != 0]

        data_narrowed.set_index('date',inplace=True)

     

        syscode_list = data2['system_code'].unique()
        if (request.POST.get('system_code00') in syscode_list):
              syscode_df = pd.DataFrame()
              for syscode in syscode_list:

                    dummy_df = data2.loc[data2['system_code'] == syscode]
                    dummy_df.drop(['system_code', 'master_total'], axis = 1, inplace = True)
                    years = data['year'].unique()
                    process_df =pd.DataFrame()
                    transposed_df = dummy_df.T
                    #print(transposed_df)
                    transposed_df = transposed_df.iloc[1:,:]
                    transposed_df.reset_index(drop = True, inplace = True)
                    transposed_df.columns = range(transposed_df.columns.size)
                    n_columns2 = transposed_df.shape[1]

                    for cols in range(0,(n_columns2)):
                        col2 = transposed_df.iloc[:, cols]
                        col2 = pd.DataFrame(col2)
                        col2.reset_index(level = 0, inplace = True)
                        process_df = process_df.append(col2)  
                    
                    process_df.set_index('index', inplace = True)

                    process_df.reset_index(drop = True, inplace =True)
                    process_df = process_df.fillna(0)
                    process_df[syscode] = process_df.iloc[:, 0:].sum(axis = 1)
                    process_df = process_df[[ syscode]]
                  
                    syscode_df = pd.concat([syscode_df,process_df], axis = 1)
                    process_df = process_df[0:0]

          
              data_narrowed.reset_index(drop=False, inplace = True)
              syscode_df = syscode_df[0:36]
              syscode_df = pd.concat([syscode_df, data_narrowed], axis = 1)
              syscode_df = syscode_df.fillna(0)
              syscode_df = syscode_df.set_index('date')
              mean = str(round(syscode_df[sysCode].mean(),2))
              for syscode in syscode_list:
                  syscode_df[syscode] =syscode_df[syscode].replace(0,syscode_df[syscode].mean() )
              forecast_df = pd.DataFrame()
              
              maxima = len(data_narrowed)
              future_dates=[syscode_df.index[-1] + DateOffset(months=x) for x in range(0,duration2)]

              
              model2=sm.tsa.statespace.SARIMAX(syscode_df[sysCode],order=(1, 1, 1),seasonal_order=(1,1,1, seasonal2), enforce_stationarity=False)
              results=model2.fit()


              future_dataset_df = pd.DataFrame(index=future_dates[1:], columns = syscode_df.columns)

              future_df = pd.concat([syscode_df, future_dataset_df])

              future_df[sysCode+'_forecast']= results.predict(start= initializer2, end= maxima + int(request.POST.get('duration2')),dynamic=True)
              future_df[future_df < 0] = 0
              future_df[[sysCode,sysCode+'_forecast']].plot(figsize=(10,8))
              pyplot.xlabel("Time")
              pyplot.ylabel("Expenditure (in Euros)")
              pyplot.title("Forecast fitting for the System Code: " + sysCode+ "   (0 level of the original plot (blue line) has been shifted with a bias of: "+ mean + " for training purposes)")
              pyplot.show()
              return render(request, 'dl_operations.html')

        else:
            return render(request, 'dl_operations.html')


      if request.POST.get('excel'):
           
            objects = AITrainerData.objects.raw( 'SELECT * FROM ai_trainer_data_scaled' )

            objects2 = AITrainerData.objects.values()
            df = pd.DataFrame(list(objects2), columns = [ 'system_code', 'year', 'january_po', 'february_po', 'march_po', 'april_po', 'may_po', 'june_po', 'july_po', 'august_po', 'september_po', 'october_po', 'november_po', 'december_po', 'january_lb', 'february_lb', 'march_lb', 'april_lb', 'may_lb', 'june_lb', 'july_lb', 'august_lb', 'september_lb', 'october_lb', 'november_lb', 'december_lb', 'january_mt', 'february_mt', 'march_mt', 'april_mt', 'may_mt', 'june_mt', 'july_mt', 'august_mt', 'september_mt', 'october_mt', 'november_mt', 'december_mt', 'completion_rate', 'master_total' ])
           
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ai_trainer_data.xlsx"'                                        
            df.to_excel(response)
            return response
      if request.POST.get('excel2'):

            objects2 = masterYearlyTotal.objects.values()
            df = pd.DataFrame(list(objects2), columns = [ 'system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total' ])
           
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="master_yearly_total.xlsx"'                                        
            df.to_excel(response)
            return response



      else:
        store = accuracyTolerance.objects.all().values()
        accuracyTot = pd.DataFrame(store)
        list_accuTot = accuracyTot.values.tolist()
        list_accuTot = np.array(list_accuTot)
        list_accuTot = list_accuTot.flatten()


        print(list_accuTot)
        accuracy = list_accuTot[1]
        
        tolerance = int(list_accuTot[2])
        return render(request, 'dl_operations.html', {'accuracy': accuracy} )
  else:
    store = accuracyTolerance.objects.all().values()
    accuracyTot = pd.DataFrame(store)
    list_accuTot = accuracyTot.values.tolist()
    list_accuTot = np.array(list_accuTot)
    list_accuTot = list_accuTot.flatten()

    print(list_accuTot)
    accuracy = list_accuTot[1]
        
    tolerance = int(list_accuTot[2])

    return render(request, 'dl_operations.html', {'accuracy': accuracy} )

def data_analysis(request):
  if(request.method == 'POST'):

    objectsall = requiredTable.objects.all().values()
    total_expenditure = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
    total_expenditure = pd.DataFrame(objectsall)
    summation = total_expenditure['total_nrc'].sum()


    if (request.POST.get('total')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = False, inplace= True)
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace(['.'], ',')
      json_records = all_objects.reset_index().to_json(orient = 'records')
      data = []
      data = json.loads(json_records)
     
      return render (request, 'data_analysis.html', {"data":data, "years": 2019, 'sort': 'Maximum NRC', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.') } )

    elif (request.POST.get('total2')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = True, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace(['.'], ',')
      data = []
      data = json.loads(json_records)
     
      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Minimum NRC', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )


    elif (request.POST.get('labour_max')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['labour_cost'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace(['.'], ',')
      data = []
      data = json.loads(json_records)
     
      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Maximum Labour Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    elif (request.POST.get('labour_min')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['labour_cost'], ascending = True, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace(['.'], ',')
      data = []
      data = json.loads(json_records)
     
      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Minimum Labour Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    
    elif (request.POST.get('material_max')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['material_cost'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)
    
      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Maximum Material Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    elif (request.POST.get('material_min')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['material_cost'], ascending = True, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)
   
      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Minimum Material Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    elif (request.POST.get('purchase_max')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['purchase_orders'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)

      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Maximum Purchase Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    elif (request.POST.get('purchase_min')):
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['purchase_orders'], ascending = True, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)

      return render (request, 'data_analysis.html', {"data":data, "years": 2019,  'sort': 'Minimum Purchase Cost', 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')} )

    

    elif(request.POST.get('details')):
      systemcode =  request.POST.get('syscode')
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)
      system = str(systemcode)
      objects = requiredTable.objects.raw('SELECT * FROM required_table WHERE system_code = %s', [systemcode])
      messages.error(request, "Overall NRC till date for system code: "+system)
      return render(request, 'data_analysis.html', {'total_nrc': objects, 'syscode': systemcode, 'data':data, "years": 2019, 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')})
    
    elif(request.POST.get('overall')):
      systemcode =  request.POST.get('syscode')
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      total_labour_cost = all_objects['labour_cost'].sum()
      total_material_cost = all_objects['material_cost'].sum()
      total_purchase_cost = all_objects['purchase_orders'].sum()
      all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']] = all_objects[['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate']].astype(str)
      all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']] = all_objects[['labour_cost', 'material_cost', 'purchase_orders', 'total_nrc']].replace('.', ',')
      data = []
      data = json.loads(json_records)
      objects = requiredTable.objects.raw('SELECT * FROM required_table WHERE system_code = %s', [systemcode])
      system = str(systemcode)
      messages.success(request, "Overall NRC till date for system code: "+system )
      return render(request, 'data_analysis.html', {'total_nrc': objects, 'syscode': systemcode, 'data':data, "years": 2019 , 'summation': str(round(summation,2)).replace(',','.'), 'lbc': str(round(total_labour_cost,2)).replace(',', '.'), 'mtc': str(round(total_material_cost,2)).replace(',', '.'), 'tpc': str(round(total_purchase_cost,2)).replace(',','.')})

    elif(request.POST.get('choose')):
      
      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      alldata = []
      alldata = json.loads(json_records)

      systemcode =  request.POST.get('syscode')
      year = int(request.POST.get('year'))
      
      objects1 = labourMaster.objects.all().values()
      objects2 = materialMaster.objects.all().values()
      objects3 = purchaseYearlyTotal.objects.all().values()
      objects4 = completionRate.objects.all().values()
      labour = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      material = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      purchase = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      comprate = pd.DataFrame(columns=['system_code', 'completion_rate'])
      labour = pd.DataFrame(objects1)
      material = pd.DataFrame(objects2)
      purchase = pd.DataFrame(objects3)
      comprate= pd.DataFrame(objects4)
      comprate = comprate.dropna(axis = 0)
      json_df = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      labour = labour.loc[((labour['system_code'] == systemcode) & (labour['year'] == year)) ]
      material = material.loc[((material['system_code'] == systemcode) & (material['year'] == year))]
      purchase = purchase.loc[((purchase['system_code'] == systemcode) & (purchase['year'] == year))]
      comprate = comprate.loc[comprate['system_code'] == systemcode]


      labour = labour.append(material)
      labour = labour.append(purchase)
      labour.reset_index(drop =True, inplace = True)
      comprate.reset_index(drop = True, inplace = True)
      json_df['system_code'] = labour['system_code']
      json_df['labour_cost'] = labour.loc[0,'total'] if len(labour) >= 1 else 0
      json_df['material_cost'] = labour.loc[1,'total'] if len(labour) >= 2 else 0
      json_df['purchase_orders'] = labour.loc[2,'total'] if len(labour) == 3 else 0
      json_df['total_nrc'] = json_df.loc[:, 'labour_cost':'purchase_orders'].sum(axis=1)
 
      json_df = json_df.iloc[:1,:]
      json_df['completion_rate'] = comprate['completion_rate']
      

      json_records = json_df.reset_index().to_json(orient = 'records')
      data = []
      data = json.loads(json_records)
      str_year = str(year)
      system = str(systemcode)
      messages.success(request, "Expenditures on system code " +system+ " for the year " +str_year )
      return render(request, 'data_analysis.html',{'syscode': systemcode, 'data': alldata, 'total_nrc': data, 'years':year , 'summation': round(summation,2)})

    elif(request.POST.get('choose1')):

      objectsall = requiredTable.objects.all().values()
      all_objects = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      all_objects = pd.DataFrame(objectsall)
      all_objects.sort_values(by = ['total_nrc'], ascending = False, inplace= True)
      json_records = all_objects.reset_index().to_json(orient = 'records')
      alldata = []
      alldata = json.loads(json_records)
      systemcode =  request.POST.get('syscode')
      year = int(request.POST.get('year'))
      month = request.POST.get('month')
      str_year = str(year)
      system = str(systemcode)
      objects1 = labourMaster.objects.all().values()
      objects2 = materialMaster.objects.all().values()
      objects3 = purchaseYearlyTotal.objects.all().values()
      objects4 = completionRate.objects.all().values()
      labour = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      labour_tot = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      labour_last = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      material = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      material_last = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      material_tot = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      purchase = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      purchase_tot = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      purchase_last = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      comprate = pd.DataFrame(columns=['system_code', 'completion_rate'])
      labour_yearly = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      purchase_yearly = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      material_yearly = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
      labour = pd.DataFrame(objects1)
      labour_tot = pd.DataFrame(objects1)
      material = pd.DataFrame(objects2)
      material_tot = pd.DataFrame(objects2)
      purchase = pd.DataFrame(objects3)
      purchase_tot = pd.DataFrame(objects3)
      comprate= pd.DataFrame(objects4)
      comprate = comprate.dropna(axis = 0)
      json_df = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      json_df2 = pd.DataFrame(columns = ['system_code', 'labour_cost', 'material_cost', 'purchase_orders', 'total_nrc', 'completion_rate'])
      labour = labour.loc[((labour['system_code'] == systemcode) & (labour['year'] == year)) ]
      material = material.loc[((material['system_code'] == systemcode) & (material['year'] == year))]
      purchase = purchase.loc[((purchase['system_code'] == systemcode) & (purchase['year'] == year))]
      comprate = comprate.loc[comprate['system_code'] == systemcode]

      if(month == 'January'):
        labour['total'] = labour.loc[:, 'january']
        material['total'] = material.loc[:, 'january']
        purchase['total'] = purchase.loc[:, 'january']
      elif(month == 'February'):
        labour['total'] = labour.loc[:, 'february']
        material['total'] = material.loc[:, 'february']
        purchase['total'] = purchase.loc[:, 'february']
      elif(month == 'March'):
        labour['total'] = labour.loc[:, 'march']
        material['total'] = material.loc[:, 'march']
        purchase['total'] = purchase.loc[:, 'march']
      elif(month == 'April'):
        labour['total'] = labour.loc[:, 'april']
        material['total'] = material.loc[:, 'april']
        purchase['total'] = purchase.loc[:, 'april']
      elif(month == 'May'):
        labour['total'] = labour.loc[:,'may']
        material['total'] = material.loc[:,'may']
        purchase['total'] = purchase.loc[:, 'may']
      elif(month == 'June'):
        labour['total'] = labour.loc[:,'june']
        material['total'] = material.loc[:, 'june']
        purchase['total'] = purchase.loc[:, 'june']
      elif(month == 'July'):
        labour['total'] = labour.loc[:, 'july']
        material['total'] = material.loc[:, 'july']
        purchase['total'] = purchase.loc[:, 'july']
      elif(month == 'August'):
        labour['total'] = labour.loc[:, 'august']
        material['total'] = material.loc[:, 'august']
        purchase['total'] = purchase.loc[:, 'august']
      elif(month == 'September'):
        labour['total'] = labour.loc[:, 'september']
        material['total'] = material.loc[:, 'september']
        purchase['total'] = purchase.loc[:, 'september']
      elif(month == 'October'):
        labour['total'] = labour.loc[:, 'october']
        material['total'] = material.loc[:,'october']
        purchase['total'] = purchase.loc[:,'october']
      elif(month == 'November'):
        labour['total'] = labour.loc[:, 'november']
        material['total'] = material.loc[:, 'november']
        purchase['total'] = purchase.loc[:, 'november']
      elif(month == 'December'):
        labour['total'] = labour.loc[:, 'december']
        material['total'] = material.loc[:, 'december']
        purchase['total'] = purchase.loc[:, 'december']
      else:
        labour['total'] = labour['total']
        material['total'] = material['total']
        purchase['total'] = purchase['total']


      labour = labour.append(material)
      labour = labour.append(purchase)
      labour.reset_index(drop =True, inplace = True)
      comprate.reset_index(drop = True, inplace = True)
      json_df['system_code'] = labour['system_code']
      json_df['labour_cost'] = labour.loc[0,'total'] if len(labour) >= 1 else 0
      json_df['material_cost'] = labour.loc[1,'total'] if len(labour) >= 2 else 0
      json_df['purchase_orders'] = labour.loc[2,'total'] if len(labour) == 3 else 0
      json_df['total_nrc'] = json_df.loc[:, 'labour_cost':'purchase_orders'].sum(axis=1)

      json_df = json_df.iloc[:1,:]
      json_df['completion_rate'] = comprate['completion_rate']

      json_records = json_df.reset_index().to_json(orient = 'records')
      data = []
      data = json.loads(json_records)

     #Processing for the costs incurred till the selected period
      year_len = year - 2019
      labour_tot = labour_tot.loc[labour_tot['system_code'] == systemcode]
      labour_tot.reset_index(drop=True, inplace = True)
      material_tot = material_tot.loc[material_tot['system_code'] == systemcode]
      material_tot.reset_index(drop = True, inplace = True)
      purchase_tot = purchase_tot.loc[purchase_tot['system_code'] == systemcode]
      purchase.reset_index(drop = True, inplace = True)
        
       
      labour_last = labour_tot.loc[labour_tot['year'] == year]
      labour_last.reset_index(drop=True, inplace = True)
      labour_tot = labour_tot.iloc[: year_len, :]

      material_last = material_tot.loc[material_tot['year'] == year]
      material_last.reset_index(drop=True, inplace = True)
      material_tot = material_tot.iloc[: year_len, :]

      purchase_last = purchase_tot.loc[purchase_tot['year'] == year]
      purchase_last.reset_index(drop=True, inplace = True)
      purchase_tot = purchase_tot.iloc[: year_len, :]
     
      
      if(month == 'January'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last['january'] if year != 2019 else labour_last['january']

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last['january'] if year != 2019 else material_last['january']

        purchase_tot = purchase_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last['january'] if year != 2019 else purchase_last['january']

      elif(month == 'February'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'february'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'february'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'february'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'february'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'february'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'february'].sum(axis = 1)

      elif(month == 'March'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'march'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'march'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'march'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'march'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'march'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'march'].sum(axis = 1)

      elif(month == 'April'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'april'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'april'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'april'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'april'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'april'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'april'].sum(axis = 1)

      elif(month == 'May'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'may'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'may'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'may'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'may'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'may'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'may'].sum(axis = 1)

      elif(month == 'June'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'june'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'june'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'june'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'june'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'june'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'june'].sum(axis = 1)

      elif(month == 'July'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'july'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'july'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'july'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'july'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'july'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'july'].sum(axis = 1)

      elif(month == 'August'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'august'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'august'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'august'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'august'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'august'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'august'].sum(axis = 1)

      elif(month == 'September'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'september'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'september'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'september'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'september'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'september'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'september'].sum(axis = 1)

      elif(month == 'October'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'october'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'october'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'october'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'october'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'october'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'october'].sum(axis = 1)

      elif(month == 'November'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'november'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'november'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'november'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'november'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'november'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'november'].sum(axis = 1)

      elif(month == 'December'):
        labour_tot = labour_tot[['system_code', 'total']]
        labour_tot = labour_tot.groupby(['system_code'], as_index = False).sum()
        labour_last['total'] =  labour_tot['total'] + labour_last.loc[:, 'january':'december'].sum(axis = 1) if year != 2019 else labour_last.loc[:, 'january':'december'].sum(axis = 1)

        material_tot = material_tot[['system_code', 'total']]
        material_tot = material_tot.groupby(['system_code'], as_index = False).sum()
        material_last['total'] =  material_tot['total'] + material_last.loc[:, 'january':'december'].sum(axis = 1) if year != 2019 else material_last.loc[:, 'january':'december'].sum(axis = 1)

        purchase_tot = labour_tot[['system_code', 'total']]
        purchase_tot = purchase_tot.groupby(['system_code'], as_index = False).sum()
        purchase_last['total'] =  purchase_tot['total'] + purchase_last.loc[:, 'january':'december'].sum(axis = 1) if year != 2019 else purchase_last.loc[:, 'january':'december'].sum(axis = 1)
      
    
      #Processing for the cost incurred over all system codes till date

  
      lb_aggregate = labourMasterAggregate.objects.all().values()
      mt_aggregate = materialMasterAggregate.objects.all().values()
      po_aggregate = purchaseAggregate.objects.all().values()
      labour_yearly = pd.DataFrame(lb_aggregate)
      material_yearly = pd.DataFrame(mt_aggregate)
      purchase_yearly = pd.DataFrame(po_aggregate)


      labour_yearly = pd.DataFrame(objects1)
      material_yearly = pd.DataFrame(objects2)
      purchase_yearly = pd.DataFrame(objects3)
      labour_yearly_sum = pd.DataFrame(objects1)
     
      
      labour_yearly_sum = labour_yearly.groupby(['year'], as_index = False).sum()
      labour_yearly_sum.reset_index(drop = True, inplace = True)
      labour_yearly_sum.drop(['id'], inplace = True, axis = 1)

      material_yearly_sum = material_yearly.groupby(['year'], as_index = False).sum()
      material_yearly_sum.reset_index(drop = True, inplace = True)
      material_yearly_sum.drop(['id'], inplace = True, axis = 1)

      purchase_yearly_sum = purchase_yearly.groupby(['year'], as_index = False).sum()
      purchase_yearly_sum.reset_index(drop = True, inplace = True)
      purchase_yearly_sum.drop(['id'], inplace = True, axis = 1)




      this_year = int(year)
      this_month = month.lower()
      lab_tot = pd.DataFrame(columns = ['total'])
      mat_tot = pd.DataFrame(columns = ['total'])
      pur_tot = pd.DataFrame(columns = ['total'])

      print(labour_yearly_sum.head())
     
      for index, rows in labour_yearly_sum.iterrows():
      
        if( int(rows['year']) != this_year and this_year != 2019 ):
          lab_tot = lab_tot.append(dict(total = rows['total']), ignore_index = True)
        else:
          lab_tot =  lab_tot.append( dict (total = rows['january': this_month].sum() ), ignore_index= True)
          break

      for index, rows in material_yearly_sum.iterrows():
      
        if( int(rows['year']) != this_year and this_year != 2019 ):
          mat_tot = mat_tot.append(dict(total = rows['total']), ignore_index = True)
        else:
          mat_tot =  mat_tot.append( dict (total = rows['january': this_month].sum() ), ignore_index= True)
          break

      for index, rows in purchase_yearly_sum.iterrows():
      
        if( int(rows['year']) != this_year and this_year != 2019 ):
          pur_tot = pur_tot.append(dict(total = rows['total']), ignore_index = True)
        else:
          pur_tot =  pur_tot.append( dict (total = rows['january': this_month].sum() ), ignore_index= True)
          break
       
      
      labour_total = lab_tot['total'].sum()
      material_total = mat_tot['total'].sum()
      purchase_total = pur_tot['total'].sum()

     
    
      labour_last = labour_last[['system_code', 'total']]
      material_last = material_last[['system_code', 'total']]
      purchase_last = purchase_last[['system_code', 'total']]

      labour_last = labour_last.append(material_last)
      labour_last = labour_last.append(purchase_last)
      labour_last.reset_index(inplace = True, drop = True)

      json_df2['system_code'] = labour_last['system_code']
      json_df2['labour_cost'] = labour_last.loc[0,'total'] if len(labour_last) >= 1 else 0
      json_df2['material_cost'] = labour_last.loc[1,'total'] if len(labour_last) >= 2 else 0
      json_df2['purchase_orders'] = labour_last.loc[2,'total'] if len(labour_last) == 3 else 0
      json_df2['total_nrc'] = json_df2.loc[:, 'labour_cost':'purchase_orders'].sum(axis=1)

      json_df2 = json_df2.iloc[:1,:]
      json_df2['completion_rate'] = comprate['completion_rate']

      json_records2 = json_df2.reset_index().to_json(orient = 'records')
      data2 = []
      data2 = json.loads(json_records2)
      

      print(labour_last)

      
      messages.success(request, "Expenditures on system code " +system+ " for the year " +str_year+ " and month of " +month)
      return render(request, 'data_analysis.html',{'syscode': systemcode, 'data': alldata, 'total_nrc2': data2, 'total_nrc': data, 'years':year , 'summation': round(summation,2), 'labour_total': round(labour_total,2), 'material_total': round(material_total,2), 'purchase_total': round(purchase_total,2)})

    else:

      return render(request, 'data_analysis.html', {'summation': round(summation,2)})
  else:
    return render(request, 'data_analysis.html')



def plotter():
        df = pd.DataFrame(columns = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december','index', 'label'])
        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        df3 = pd.DataFrame()
        x_data1 = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']  
        x_data2 = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        x_data3 = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        graph_data = graphData.objects.all().values()
        df = pd.DataFrame(graph_data)
        df1 = df.loc[df['index'] == 1] if len(df) != 0 and 1 in df['index'].values else df1
        df2 = df.loc[df['index'] == 2] if len(df) != 0 and 2 in df['index'].values else df2
        df3 = df.loc[df['index'] == 3] if len(df) != 0  and 3 in df['index'].values else df3
        label_1 = str(df1['label'].values) if len(df1) !=0 else 'none'
        label_2 = str(df2['label'].values) if len(df2) !=0 else 'none'
        label_3 = str(df3['label'].values) if len(df3) !=0 else 'none'
        if(len(df1)!= 0 ):
            df1.drop(['id', 'index', 'label'], inplace = True, axis = 1)
        if(len(df2)!= 0 ):
            df2.drop(['id', 'index', 'label'], inplace = True, axis = 1)
        if(len(df3)!= 0 ):
            df3.drop(['id', 'index', 'label'], inplace = True, axis = 1)
        yy_data1 = df1.values.flatten()
        yy_data2 = df2.values.flatten()
        yy_data3 = df3.values.flatten()
        df = df[0:0]
        plot_div = plot([Scatter(x=x_data1, y=yy_data1,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='blue')], 
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="")
       
                
        plot_div2 = plot([Scatter(x=x_data2, y=yy_data2,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='blue')],
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="")

        plot_div3 = plot([Scatter(x=x_data3, y=yy_data3,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='blue')],
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="")


        plot_div = plot_div if len(df1) !=0 else 'none'
        label_1 = label_1 if len(df1)  != 0 else 'none'

        plot_div2 = plot_div2 if len(df2) !=0 else 'none'
        label_2 = label_2 if len(df2)  != 0 else 'none'

        plot_div3 = plot_div3 if len(df3) !=0 else 'none'
        label_3 = label_3 if len(df3)  != 0 else 'none'

        return plot_div, plot_div2, plot_div3, label_1, label_2, label_3

def plotter_2():
        df = pd.DataFrame(columns = ['index', 'systemcode', 'year', 'label', 'total'])
        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        df3 = pd.DataFrame()
        yearly_graph_data = yearlyGraphData.objects.all().values()
        df = pd.DataFrame(yearly_graph_data)

        df1 = df.loc[df['index'] == 1] if len(df) != 0 and 1 in df['index'].values else df1
        df2 = df.loc[df['index'] == 2] if len(df) != 0 and 2 in df['index'].values else df2
        df3 = df.loc[df['index'] == 3] if len(df) != 0  and 3 in df['index'].values else df3

        if(len(df1)!= 0 ):
            df1.drop(['id', 'index'], inplace = True, axis = 1)
        if(len(df2)!= 0 ):
            df2.drop(['id', 'index'], inplace = True, axis = 1)
        if(len(df3)!= 0 ):
            df3.drop(['id', 'index'], inplace = True, axis = 1)
        df1.reset_index(drop = True, inplace = True)
        df2.reset_index(drop = True, inplace = True)
        df3.reset_index(drop = True, inplace = True)
       
        label2_1 = str(df1['label'][0]) if len(df1) !=0 else 'none'
        label2_2 = str(df2['label'][0]) if len(df2) != 0 else 'none'
        label2_3 = str(df3['label'][0]) if len(df3) != 0 else 'none'
        xx_data1 = df1['year'].tolist() if len(df1)!= 0 else [0,0,0,0]
        xx_data2 = df2['year'].tolist() if len(df2)!= 0 else [0,0,0,0]
        xx_data3 = df3['year'].tolist() if len(df3)!= 0 else [0,0,0,0]
        yy_data1 = df1['total'].tolist() if len(df1)!= 0 else [0,0,0,0]
        yy_data2 = df2['total'].tolist() if len(df2)!= 0 else [0,0,0,0]
        yy_data3 = df3['total'].tolist() if len(df3)!= 0 else [0,0,0,0]
        df = df[0:0]
        plot2_div = plot([Scatter(x=xx_data1, y=yy_data1,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='green', 
                                )],
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="")
                
        plot2_div2 = plot([Scatter(x=xx_data2, y=yy_data2,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='green',
                                 )], 
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="")

        plot2_div3 = plot([Scatter(x=xx_data3, y=yy_data3,
                                  mode='lines+markers', name='test',
                                  opacity=0.8, marker_color='green',
                                 )],
                        output_type='div', include_plotlyjs= False, show_link=False, link_text="" )

        plot2_div = plot2_div if len(df1) !=0 else 'none'
        label2_1 = label2_1 if len(df1)  != 0 else 'none'

        plot2_div2 = plot2_div2 if len(df2) !=0 else 'none'
        label2_2 = label2_2 if len(df2)  != 0 else 'none'

        plot2_div3 = plot2_div3 if len(df3) !=0 else 'none'
        label2_3 = label2_3 if len(df3)  != 0 else 'none'


        return plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3


        

def graphical_data_analysis(request):
 
  if(request.method == 'POST'):
    
    
    if(request.POST.get("cu_lab_cost") and request.POST.get("syscode") and request.POST.get("cu_lab_year") ):

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                cu_lab_year = int(request.POST.get("cu_lab_year"))
                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == cu_lab_year))]
                print(labour_df)
                y_data1 = labour_df.loc[:,'january':'december'].values.flatten()
                y_data1 = np.cumsum(y_data1)
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Cumulative Labour Cost (in euros) for the year " + str(cu_lab_year)  + " for the system code: " + syscode
                model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                cu_lab_year = int(request.POST.get("cu_lab_year"))

                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == cu_lab_year))]
                print(labour_df)
                y_data2 = labour_df.loc[:,'january':'december'].values.flatten()
                y_data2 = np.cumsum(y_data2)
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Cumulative Labour Cost (in euros) for the year " + str(cu_lab_year) + " for the system code: " + syscode
                model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

    
                syscode = request.POST.get("syscode")
                cu_lab_year = int(request.POST.get("cu_lab_year"))

                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == cu_lab_year))]
                print(labour_df)
                y_data3 = labour_df.loc[:,'january':'december'].values.flatten()
                y_data3 = np.cumsum(y_data3)
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Cumulative Labour Cost (in euros) for the year " + str(cu_lab_year) + " for the system code: " + syscode
                model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")


    elif(request.POST.get("lab_cost") and request.POST.get("syscode") and request.POST.get("lab_year") ):

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                lab_year = int(request.POST.get("lab_year"))
                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == lab_year))]
                print(labour_df)
                y_data1 = labour_df.loc[:,'january':'december'].values.flatten()
               
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Monthly Labour Cost (in euros) for the year " + str(lab_year) + " for the system code: " + syscode
                model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                lab_year = int(request.POST.get("lab_year"))

                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == lab_year))]
                print(labour_df)
                y_data2 = labour_df.loc[:,'january':'december'].values.flatten()
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Monthly Labour Cost (in euros) for the year " + str(lab_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

    
                syscode = request.POST.get("syscode")
                lab_year = int(request.POST.get("lab_year"))

                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode) & (labour_df['year'] == lab_year))]
                print(labour_df)
                y_data3 = labour_df.loc[:,'january':'december'].values.flatten()
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Monthly Labour Cost (in euros) for the year " + str(lab_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")


    elif(request.POST.get("cu_mat_cost") and request.POST.get("syscode") and request.POST.get("cu_mat_year") ):

        

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                cu_mat_year = int(request.POST.get("cu_mat_year"))
                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == cu_mat_year))]
                print(material_df)
                y_data1 = material_df.loc[:,'january':'december'].values.flatten()
                y_data1 = np.cumsum(y_data1)
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Cumulative Material Cost (in euros) for the year " + str(cu_mat_year) + " for the system code: " + syscode
                model.save()
                
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                cu_mat_year = int(request.POST.get("cu_mat_year"))

                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == cu_mat_year))]
                print(material_df)
                y_data2 = material_df.loc[:,'january':'december'].values.flatten()
                y_data2 = np.cumsum(y_data2)
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Cumulative Material Cost (in euros) for the year " + str(cu_mat_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

    
                syscode = request.POST.get("syscode")
                cu_mat_year = int(request.POST.get("cu_mat_year"))

                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == cu_mat_year))]
                print(material_df)
                y_data3 = material_df.loc[:,'january':'december'].values.flatten()
                y_data3 = np.cumsum(y_data3)
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Cumulative Material Cost (in euros) for the year " + str(cu_mat_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")


    elif(request.POST.get("mat_cost") and request.POST.get("syscode") and request.POST.get("mat_year") ):


        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                mat_year = int(request.POST.get("mat_year"))
                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == mat_year))]
                print(material_df)
                y_data1 = material_df.loc[:,'january':'december'].values.flatten()
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Monthly Material Cost (in euros) for the year " + str(mat_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                mat_year = int(request.POST.get("mat_year"))

                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == mat_year))]
                print(material_df)
                y_data2 = material_df.loc[:,'january':'december'].values.flatten()
               
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Monthly Material Cost (in euros) for the year " + str(mat_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

    
                syscode = request.POST.get("syscode")
                mat_year = int(request.POST.get("mat_year"))

                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                
                material_df = material_df.loc[((material_df['system_code'] == syscode) & (material_df['year'] == mat_year))]
                print(material_df)
                y_data3 = material_df.loc[:,'january':'december'].values.flatten()
            
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Monthly Material Cost (in euros) for the year " + str(mat_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")




    elif(request.POST.get("cu_pur_cost") and request.POST.get("syscode") and request.POST.get("cu_pur_year") ):

        
        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                cu_pur_year = int(request.POST.get("cu_pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == cu_pur_year))]
                print(purchase_df)
                y_data1 = purchase_df.loc[:,'january':'december'].values.flatten()
                y_data1 = np.cumsum(y_data1)
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Cumulative Purchase Cost (in euros) for the year " + str(cu_pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                cu_pur_year = int(request.POST.get("cu_pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == cu_pur_year))]
                print(purchase_df)
                y_data2 = purchase_df.loc[:,'january':'december'].values.flatten()
                y_data2 = np.cumsum(y_data2)
              
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Cumulative Purchase Cost (in euros) for the year " + str(cu_pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

                syscode = request.POST.get("syscode")
                cu_pur_year = int(request.POST.get("cu_pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == cu_pur_year))]
                print(purchase_df)
                y_data3 = purchase_df.loc[:,'january':'december'].values.flatten()
                y_data3 = np.cumsum(y_data3)
              
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Cumulative Purchase Cost (in euros) for the year " + str(cu_pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")

    elif(request.POST.get("pur_cost") and request.POST.get("syscode") and request.POST.get("pur_year") ):


        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                pur_year = int(request.POST.get("pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == pur_year))]
                print(purchase_df)
                y_data1 = purchase_df.loc[:,'january':'december'].values.flatten()
               
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0]  if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Monthly Purchase Cost (in euros) for the year " + str(pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                pur_year = int(request.POST.get("pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == pur_year))]
                print(purchase_df)
                y_data2 = purchase_df.loc[:,'january':'december'].values.flatten()
               
              
                print(y_data2)
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Monthly Purchase Cost (in euros) for the year " + str(pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

                syscode = request.POST.get("syscode")
                pur_year = int(request.POST.get("pur_year"))
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode) & (purchase_df['year'] == pur_year))]
                print(purchase_df)
                y_data3 = purchase_df.loc[:,'january':'december'].values.flatten()
                print(y_data3)
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1]  if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8]  if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Monthly Purchase Cost (in euros) for the year " + str(pur_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label_1.strip("[']").rstrip("'")
                label2_2 = label_2.strip("[']").rstrip("'")
                label2_3 = label_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        else:
         messages.success(request, "Please select one of the plots to render the data")
         return render(request, "graphical_data_analysis.html")
    
    elif(request.POST.get("y_lab_cost") and request.POST.get("syscode")):

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode))]
                labour_df = labour_df[['system_code', 'year', 'total']]
                labour_df.reset_index(drop = True, inplace =True)
                
                
                yearlyGraphData.objects.filter(index =1).delete()
                graphData.objects.filter(index = 1).delete()
                
                print(labour_df)
                for index, rows in labour_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 1
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly labour cost (in euros)" + " for the system code: " + syscode
                  model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})
        
        elif(request.POST.get("plot") == "plot_2"):
                syscode = request.POST.get("syscode")
                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode))]
                labour_df = labour_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =2).delete()
                graphData.objects.filter(index = 2).delete()
               
                for index, rows in labour_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 2
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly labour cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot2_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label2_2, 'label3': this_label3})

        elif(request.POST.get("plot") == "plot_3"):
                syscode = request.POST.get("syscode")
                objectsall = labourMaster.objects.all().values()
                labour_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                labour_df = pd.DataFrame(objectsall)
                labour_df = labour_df.loc[((labour_df['system_code'] == syscode))]
                labour_df = labour_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =3).delete()
                graphData.objects.filter(index = 3).delete()
                
                for index, rows in labour_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 3
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly labour cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot2_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label2_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")

    elif(request.POST.get("y_mat_cost") and request.POST.get("syscode")):

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                material_df = material_df.loc[((material_df['system_code'] == syscode))]
                material_df = material_df[['system_code', 'year', 'total']]
                material_df.reset_index(drop = True, inplace =True)
                
                
                yearlyGraphData.objects.filter(index =1).delete()
                graphData.objects.filter(index = 1).delete()
                
                print(material_df)
                for index, rows in material_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 1
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Material cost (in euros)" + " for the system code: " + syscode
                  model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})
        
        elif(request.POST.get("plot") == "plot_2"):
                syscode = request.POST.get("syscode")
                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                material_df = material_df.loc[((material_df['system_code'] == syscode))]
                material_df = material_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =2).delete()
                graphData.objects.filter(index = 2).delete()
               
                for index, rows in material_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 2
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Material cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot2_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label2_2, 'label3': this_label3})

        elif(request.POST.get("plot") == "plot_3"):
                syscode = request.POST.get("syscode")
                objectsall = materialMaster.objects.all().values()
                material_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                material_df = pd.DataFrame(objectsall)
                material_df = material_df.loc[((material_df['system_code'] == syscode))]
                material_df = material_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =3).delete()
                graphData.objects.filter(index = 3).delete()
                
                for index, rows in material_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 3
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Material cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot2_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label2_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")

    elif(request.POST.get("y_pur_cost") and request.POST.get("syscode")):

        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode))]
                purchase_df = purchase_df[['system_code', 'year', 'total']]
                purchase_df.reset_index(drop = True, inplace =True)
                
                
                yearlyGraphData.objects.filter(index =1).delete()
                graphData.objects.filter(index = 1).delete()
                
                print(purchase_df)
                for index, rows in purchase_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 1
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Purchase cost (in euros)" + " for the system code: " + syscode
                  model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})
        
        elif(request.POST.get("plot") == "plot_2"):
                syscode = request.POST.get("syscode")
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode))]
                purchase_df = purchase_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =2).delete()
                graphData.objects.filter(index = 2).delete()
               
                for index, rows in purchase_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 2
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Purchase cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot2_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label2_2, 'label3': this_label3})

        elif(request.POST.get("plot") == "plot_3"):
                syscode = request.POST.get("syscode")
                objectsall = purchaseYearlyTotal.objects.all().values()
                purchase_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'total'])
                purchase_df = pd.DataFrame(objectsall)
                purchase_df = purchase_df.loc[((purchase_df['system_code'] == syscode))]
                purchase_df = purchase_df[['system_code', 'year', 'total']]

                yearlyGraphData.objects.filter(index =3).delete()
                graphData.objects.filter(index = 3).delete()
                
                for index, rows in purchase_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 3
                  model.systemcode = rows['system_code']
                  model.year = rows['year']
                  model.total = rows['total']
                  model.label = "Yearly Purchase cost (in euros)" + " for the system code: " + syscode
                  model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot2_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label2_3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")



    elif(request.POST.get("yearly_total") and request.POST.get("syscode") and request.POST.get("y_total") ):


        if(request.POST.get("plot") == "plot_1"):
                syscode = request.POST.get("syscode")
                master_year = int(request.POST.get("yearly_total"))
                objectsall = masterYearlyTotal.objects.all().values()
                master_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                master_df = pd.DataFrame(objectsall)
                
                master_df = master_df.loc[((master_df['system_code'] == syscode) & (master_df['year'] == master_year))]
                print(master_df)
                y_data1 = master_df.loc[:,'january':'december'].values.flatten()
              
                print(y_data1)
                
                graphData.objects.filter(index =1).delete()
                yearlyGraphData.objects.filter(index=1).delete()
                model = graphData()
                model.index = 1
                model.january = y_data1[0] if len(y_data1) != 0 else 0
                model.february = y_data1[1] if len(y_data1) != 0 else 0
                model.march = y_data1[2] if len(y_data1) != 0 else 0
                model.april = y_data1[3] if len(y_data1) != 0 else 0
                model.may = y_data1[4] if len(y_data1) != 0 else 0
                model.june = y_data1[5] if len(y_data1) != 0 else 0
                model.july = y_data1[6] if len(y_data1) != 0 else 0
                model.august = y_data1[7] if len(y_data1) != 0 else 0
                model.september = y_data1[8] if len(y_data1) != 0 else 0
                model.october = y_data1[9] if len(y_data1) != 0 else 0
                model.november = y_data1[10] if len(y_data1) != 0 else 0
                model.december = y_data1[11] if len(y_data1) != 0 else 0
                model.label = "Monthly Overall Expenditure (in euros) for the year " + str(master_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': plot_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label_1, 'label2': this_label2, 'label3': this_label3})
                
        elif(request.POST.get("plot") == "plot_2"):
    
                syscode = request.POST.get("syscode")
                master_year = int(request.POST.get("yearly_total"))
                objectsall = masterYearlyTotal.objects.all().values()
                master_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                master_df = pd.DataFrame(objectsall)
                
                master_df = master_df.loc[((master_df['system_code'] == syscode) & (master_df['year'] == master_year))]
                print(master_df)
                y_data2 = master_df.loc[:,'january':'december'].values.flatten()
              
             
                graphData.objects.filter(index =2).delete()
                yearlyGraphData.objects.filter(index=2).delete()
                model = graphData()
                model.index = 2
                model.january = y_data2[0] if len(y_data2) != 0 else 0
                model.february = y_data2[1] if len(y_data2) != 0 else 0
                model.march = y_data2[2] if len(y_data2) != 0 else 0
                model.april = y_data2[3] if len(y_data2) != 0 else 0
                model.may = y_data2[4] if len(y_data2) != 0 else 0
                model.june = y_data2[5] if len(y_data2) != 0 else 0
                model.july = y_data2[6] if len(y_data2) != 0 else 0
                model.august = y_data2[7] if len(y_data2) != 0 else 0
                model.september = y_data2[8] if len(y_data2) != 0 else 0
                model.october = y_data2[9] if len(y_data2) != 0 else 0
                model.november = y_data2[10] if len(y_data2) != 0 else 0
                model.december = y_data2[11] if len(y_data2) != 0 else 0
                model.label = "Monthly Overall Expenditure (in euros) for the year " + str(master_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': plot_div2, 'plot_div3': this_div3, 'label1': this_label1, 'label2': label_2, 'label3': this_label3})
                           
   
        elif(request.POST.get("plot") == "plot_3"):

    
                syscode = request.POST.get("syscode")
                master_year = int(request.POST.get("yearly_total"))
                objectsall = masterYearlyTotal.objects.all().values()
                master_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                master_df = pd.DataFrame(objectsall)
                
                master_df = master_df.loc[((master_df['system_code'] == syscode) & (master_df['year'] == master_year))]
                print(master_df)
                y_data3 = master_df.loc[:,'january':'december'].values.flatten()
              
                 
                graphData.objects.filter(index =3).delete()
                yearlyGraphData.objects.filter(index=3).delete()
                model = graphData()
                model.index = 3
                model.january = y_data3[0] if len(y_data3) != 0 else 0
                model.february = y_data3[1] if len(y_data3) != 0 else 0
                model.march = y_data3[2] if len(y_data3) != 0 else 0
                model.april = y_data3[3] if len(y_data3) != 0 else 0
                model.may = y_data3[4] if len(y_data3) != 0 else 0
                model.june = y_data3[5] if len(y_data3) != 0 else 0
                model.july = y_data3[6] if len(y_data3) != 0 else 0
                model.august = y_data3[7] if len(y_data3) != 0 else 0
                model.september = y_data3[8] if len(y_data3) != 0 else 0
                model.october = y_data3[9] if len(y_data3) != 0 else 0
                model.november = y_data3[10] if len(y_data3) != 0 else 0
                model.december = y_data3[11] if len(y_data3) != 0 else 0
                model.label = "Monthly Overall Expenditure (in euros) for the year " + str(master_year) + " for the system code: " + syscode
                model.save()
                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div1 = plot_div if plot_div != 'none' else plot2_div
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_label1 = label_1 if label_1 != 'none' else label2_1
                this_label2 = label_2 if label_2 != 'none' else label2_2
                return render(request, "graphical_data_analysis.html", context={ 'systemcode':syscode,'plot_div': this_div1, 'plot_div2': this_div2, 'plot_div3': plot_div3, 'label1': this_label1, 'label2': this_label2, 'label3': label_3})
        
        else:
         
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")

    elif(request.POST.get("overall_cost") ):

        if(request.POST.get("plot") == "plot_1"):
                
              
                objectsall = masterYearlyTotal.objects.all().values()
                overall_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                overall_df = pd.DataFrame(objectsall)
                overall_df = overall_df[['year',  'master_total']]
                overall_df = overall_df.groupby(['year'], as_index = False).sum()
                overall_df.reset_index(drop = True, inplace =True)

                yearlyGraphData.objects.filter(index =1).delete()
                graphData.objects.filter(index = 1).delete()
                
                print(overall_df)
                previous_year = 0
                for index, rows in overall_df.iterrows():

                  model = yearlyGraphData()
                  model.index = 1
                  model.systemcode = 'none'
                  model.year = rows['year'] 
                  model.total = rows['master_total']
                  model.label = "Yearly Overall cost (in euros)"
                  model.save()
                 

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})
        
        elif(request.POST.get("plot") == "plot_2"):
              
                objectsall = masterYearlyTotal.objects.all().values()
                overall_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                overall_df = pd.DataFrame(objectsall)
                overall_df = overall_df[['year', 'master_total']]
                overall_df = overall_df.groupby(['year'], as_index = False).sum()
                overall_df.reset_index(drop = True, inplace =True)

                yearlyGraphData.objects.filter(index =2).delete()
                graphData.objects.filter(index = 2).delete()
                
                print(overall_df)
                for index, rows in overall_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 2
                  model.systemcode = 'none'
                  model.year = rows['year']
                  model.total = rows['master_total']
                  model.label = "Yearly Overall cost (in euros)"
                  model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})

        elif(request.POST.get("plot") == "plot_3"):
               
                objectsall = masterYearlyTotal.objects.all().values()
                overall_df = pd.DataFrame(columns = ['system_code', 'year', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december', 'master_total'])
                overall_df = pd.DataFrame(objectsall)
                overall_df = overall_df[['year', 'master_total']]
                overall_df = overall_df.groupby(['year'], as_index = False).sum()
                overall_df.reset_index(drop = True, inplace =True)

                yearlyGraphData.objects.filter(index =3).delete()
                graphData.objects.filter(index = 3).delete()
                
                print(overall_df)
                for index, rows in overall_df.iterrows():
                  model = yearlyGraphData()
                  model.index = 3
                  model.systemcode = 'none'
                  model.year = rows['year']
                  model.total = rows['master_total']
                  model.label = "Yearly Overall cost (in euros)"
                  model.save()

                plot_div, plot_div2, plot_div3, label_1, label_2, label_3 = plotter()    
                plot2_div, plot2_div2, plot2_div3, label2_1, label2_2, label2_3 = plotter_2()
                label_1 = label_1.strip("[']").rstrip("'")
                label_2 = label_2.strip("[']").rstrip("'")
                label_3 = label_3.strip("[']").rstrip("'")
                label2_1 = label2_1.strip("[']").rstrip("'")
                label2_2 = label2_2.strip("[']").rstrip("'")
                label2_3 = label2_3.strip("[']").rstrip("'")
                this_div2 = plot_div2 if plot_div2 != 'none' else plot2_div2
                this_div3 = plot_div3 if plot_div3 != 'none' else plot2_div3
                this_label2 = label_2 if label_2 != 'none' else label2_2
                this_label3 = label_3 if label_3 != 'none' else label2_3
                return render(request, "graphical_data_analysis.html", context={ 'plot_div': plot2_div, 'plot_div2': this_div2, 'plot_div3': this_div3, 'label1': label2_1, 'label2': this_label2, 'label3': this_label3})
        else:
          messages.success(request, "Please select one of the plots to render the data")
          return render(request, "graphical_data_analysis.html")


    else:
    
       return render(request, "graphical_data_analysis.html")
    

  else:
      return render(request, "graphical_data_analysis.html")
